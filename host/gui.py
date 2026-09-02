#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
图形界面版：工控机自动填入数值工具（Tkinter，Python 自带，无需额外安装界面库）。
复用 run.py 的核心函数（串口/表格/步骤引擎/断点）。

运行：
    python gui.py

功能：
    - 串口选择/刷新、Pico 在线测试
    - Excel/CSV 表格选择与步骤列校验
    - 「步骤流程」可视化编辑：点击进入界面 → 填入数值 → 返回主界面 → 进入另一界面…
      支持 click / dblclick / press_at / move / scroll / key / input / sleep / home /
      confirm / button 等步骤类型，可增删、排序、动态表单编辑
    - 「自定义按钮」：把一段常用步骤命名单成一个按钮，可插入到流程，也可在其下方快捷区一键执行
    - 坐标校准（移动光标 + 人工确认 + 像素微调，写回配置）
    - 自动填数：进度条、当前行摘要、每行人工确认后继续
    - Dry Run 预览指令序列、断点续传、日志输出
"""
import copy
import json
import os
import queue
import shutil
import sys
import threading
import time
import traceback

import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog, scrolledtext

HERE = os.path.dirname(os.path.abspath(__file__))
run = None  # 延迟导入 run.py（缺依赖时给出友好提示，见 main()）

# 步骤类型及其在表单里需要填的字段（用于启用/禁用）
STEP_TYPES = [
    "click", "dblclick", "press_at", "move", "scroll",
    "key", "input", "sleep", "home", "confirm", "button",
]
STEP_TYPE_LABELS = {
    "click": "单击", "dblclick": "双击", "press_at": "长按", "move": "移动",
    "scroll": "滚轮", "key": "按键", "input": "输入数值", "sleep": "延时",
    "home": "归零", "confirm": "人工确认", "button": "自定义按钮",
}

# 触发鼠标定位的步骤（校准用）
MOUSE_STEP_TYPES = ("click", "dblclick", "move", "press_at")
# 每个步骤类型启用的表单字段
FORM_FIELDS_BY_TYPE = {
    "click":      ("name", "x", "y", "delay"),
    "dblclick":   ("name", "x", "y", "delay"),
    "press_at":   ("name", "x", "y", "ms", "delay"),
    "move":       ("name", "x", "y", "delay"),
    "scroll":     ("name", "delta", "delay"),
    "key":        ("name", "key", "times", "delay"),
    "input":      ("name", "col", "value", "pre", "round", "enter", "delay", "prefix", "suffix", "default"),
    "sleep":      ("name", "ms"),
    "home":       ("name",),
    "confirm":    ("name",),
    "button":     ("name",),
}


def _ensure_run():
    global run
    try:
        import run as _run_mod
    except SystemExit:
        return False
    run = _run_mod
    return True


def _dump_crash(title, exc, val, tb, root=None):
    """把 GUI 运行时异常写进 host/gui_error.log，方便排查（界面退出无痕时也能看到）。"""
    try:
        msg = "".join(traceback.format_exception(exc, val, tb))
        with open(os.path.join(HERE, "gui_error.log"), "a", encoding="utf-8") as f:
            f.write("%s [%s]\n%s\n" % (time.strftime("%Y-%m-%d %H:%M:%S"), title, msg))
    except Exception:
        pass
    try:
        print(msg, file=sys.stderr)
    except Exception:
        pass


def _step_detail(s):
    """生成步骤的人类可读摘要（用于列表详情列）。"""
    typ = s.get("type")
    if typ == "click":
        return "点击 (%s,%s)" % (s.get("x"), s.get("y"))
    if typ == "dblclick":
        return "双击 (%s,%s)" % (s.get("x"), s.get("y"))
    if typ == "move":
        return "移动 (%s,%s)" % (s.get("x"), s.get("y"))
    if typ == "press_at":
        return "长按 (%s,%s) %sms" % (s.get("x"), s.get("y"), s.get("ms", 500))
    if typ in ("scroll", "wheel"):
        d = int(s.get("delta", -120))
        return "滚轮%s" % ("上" if d > 0 else "下")
    if typ == "key":
        t = max(1, int(s.get("times", 1)))
        return "%s×%d" % (str(s.get("key", "")).upper(), t)
    if typ == "input":
        src = s.get("col") if s.get("col") else s.get("value")
        pre = s.get("prefix") or ""
        suf = s.get("suffix") or ""
        extra = "（回车）" if s.get("enter") else ""
        return "输入 %s%s%s%s" % (pre, src, suf, extra)
    if typ == "sleep":
        return "等待 %sms" % s.get("ms", 0)
    if typ == "home":
        return "光标归零到屏幕左上角"
    if typ == "confirm":
        return "等待人工确认"
    if typ == "button":
        return "按钮：%s" % s.get("name")
    return typ or "?"


# ---------------- 线程与主线程间的交互代理 ----------------
class GuiInteract:
    """worker 线程不允许直接操作 Tk，所有需要用户输入的地方都走这里。"""
    def __init__(self, outbox, stop_event):
        self.outbox = outbox
        self.stop_event = stop_event
        self._results = {}
        self._event = threading.Event()
        self._next_id = 0

    def ask_yesno(self, title, prompt):
        req = self._new_req("ask_yesno", title, prompt)
        return self._wait(req) if req is not None else False

    def ask_string(self, title, prompt, init=""):
        req = self._new_req("ask_string", title, prompt, init)
        return self._wait(req)  # 取消时返回 None

    def _new_req(self, kind, title, prompt, init=""):
        if self.stop_event.is_set():
            return None
        self._next_id += 1
        req = {"id": self._next_id, "kind": kind, "title": title,
               "prompt": prompt, "init": init}
        self.outbox.put(("interact", req))
        return req

    def _wait(self, req):
        while not self._event.wait(0.2):
            self._event.clear()
            if self.stop_event.is_set():
                try:
                    self._results.pop(req["id"])
                except KeyError:
                    pass
                raise InterruptedError("用户停止了任务")
        self._event.clear()
        return self._results.pop(req["id"], None)

    def resolve(self, req, value):
        self._results[req["id"]] = value
        self._event.set()


# ---------------- 主界面 ----------------

# 正在打开但调用方已放弃/超时的连接：交给后台线程回收，防止“迟到打开”占着串口
_pending_links = []
_reaper_started = False


def _reap_late_links():
    """回收『打开超时后才成功』的迟到串口连接，防止占着端口不放。"""
    while True:
        time.sleep(1.0)
        try:
            for entry in list(_pending_links):
                t, holder = entry
                if t.is_alive():
                    continue
                link = holder.get("link")
                if link is not None:
                    try:
                        link.close()
                    except Exception:
                        pass
                    holder["link"] = None
                try:
                    _pending_links.remove(entry)
                except ValueError:
                    pass
        except Exception:
            pass


def _start_link_reaper():
    global _reaper_started
    if _reaper_started:
        return
    _reaper_started = True
    threading.Thread(target=_reap_late_links, daemon=True).start()


def _open_link_bounded(cfg, timeout=3.0):
    """有界打开串口：某些情况下 serial.Serial 打开会一直卡（端口被占用/设备拔出）。
    放在独立线程里 join 限时，超时就抛出 TimeoutError，保证『停止/重试』总能恢复。
    超时后若线程才打开成功，会由后台回收线程自动 close，避免留下幽灵连接。"""
    res = {}
    holder = {"link": None}
    def do():
        try:
            holder["link"] = run.open_link(cfg)
        except Exception as e:
            res["err"] = e
    t = threading.Thread(target=do, daemon=True)
    entry = (t, holder)
    _pending_links.append(entry)
    _start_link_reaper()
    t.start()
    t.join(timeout)
    if t.is_alive():
        _desc = ("%s:%s" % (cfg.get("wifi_ip"), cfg.get("wifi_port", 8080))
                 if (cfg.get("transport") or "serial") == "wifi" else cfg.get("com_port"))
        raise TimeoutError("打开连接 %s 超时（端口可能被占用或设备已拔出/未上电）" % _desc)
    if "err" in res:
        try:
            _pending_links.remove(entry)
        except ValueError:
            pass
        raise res["err"]
    link = holder["link"]
    try:
        _pending_links.remove(entry)
    except ValueError:
        pass
    return link


def _interruptible_sleep(ms, stop_event):
    """把延时拆成 50ms 小段，期间检查 stop_event，停止时立即返回 False。"""
    end = time.time() + max(0, ms) / 1000.0
    while time.time() < end:
        if stop_event.is_set():
            return False
        time.sleep(min(0.05, max(0.0, end - time.time())))
    return True


# 鼠标类步骤：F8 定位可以把坐标写进这些步骤
_MOUSE_STEP_SET = frozenset(("click", "dblclick", "move", "press_at"))


class App:
    def __init__(self, root):
        self.root = root
        root.title("工控机自动填入数值工具")
        # 高度需容纳：连接面板（含新增「无线控制」3 行约 100px）+ 运行时才映射的
        # 「按钮快捷执行/进度/日志」三块（合计约 520px）。沿用 820 会挤掉日志区。
        root.geometry("1400x1000")
        root.minsize(1100, 760)

        self.queue = queue.Queue()
        self.stop_event = threading.Event()
        self.confirm_event = threading.Event()
        self.worker = None
        self.cfg_path = os.path.join(HERE, "config.json")
        self.worker_cfg = None
        self.interact = None
        self.ui = {}              # 各控件引用
        self.running = False
        self.dirty = False        # 有未保存的修改（步骤/坐标/按钮）
        self.current_link = None  # worker/检测线程当前持有的串口连接（停止时强制释放用）
        self._loc_link = None     # 坐标定位期间复用的持久连接（避免每次点击都重开串口触发板子复位）
        self.device_type = "esp32"  # 当前选中的设备：esp32 / pico

        # 步骤流程对象：key -> 步骤列表。键 "__flow__" 是主流程，"btn:<名>" 是自定义按钮。
        self.objects = {"__flow__": []}
        self.object_labels = {"__flow__": "主流程"}
        self.current_obj = "__flow__"
        self._col_choices = []      # 载入表格后的表头列，供「表格列」下拉使用

        self._build_ui()
        self._load_default_config()
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)
        self.root.after(100, self._poll_queue)
        # 全局快捷键：F8 记录坐标 / F9 开始 / F10 停止 / 空格·回车确认 / Ctrl+S 保存
        self.root.bind_all("<F9>", self._shortcut_start)
        self.root.bind_all("<F10>", self._shortcut_stop)
        self.root.bind_all("<space>", self._shortcut_confirm)
        self.root.bind_all("<Return>", self._shortcut_confirm)
        self.root.bind_all("<Control-s>", self._key_save)
        self.root.bind_all("<Control-S>", self._key_save)
        self.log("快捷键：F8 坐标 / F9 开始 / F10 停止 / 空格·回车确认当前行 / Ctrl+S 保存")

    # ---------------- 界面搭建 ----------------
    def _setup_style(self):
        """统一主题色 / 字体 / 留白（纯 Tkinter + ttk，无需第三方库）。
        字体优先选『微软雅黑』这类中英文都清晰的字体，全界面统一，避免混排加粗/发虚。"""
        self.FONT = "TkDefaultFont"
        try:
            import tkinter.font as tkfont
            fams = set(tkfont.families(self.root))
            for cand in ("Microsoft YaHei UI", "Microsoft YaHei", "Segoe UI", "Noto Sans CJK SC"):
                if cand in fams:
                    self.FONT = cand
                    break
            # 统一的命名字体：常规 10pt；标题类才加粗
            for name, size, weight in (
                ("TkDefaultFont", 10, "normal"),
                ("TkTextFont", 10, "normal"),
                ("TkMenuFont", 10, "normal"),
                ("TkHeadingFont", 10, "bold"),
                ("TkSmallCaptionFont", 9, "normal"),
                ("TkIconFont", 9, "normal"),
                ("TkTooltipFont", 9, "normal"),
            ):
                try:
                    tkfont.nametofont(name).configure(family=self.FONT, size=size, weight=weight)
                except Exception:
                    pass
            style = ttk.Style(self.root)
            try:
                style.theme_use("clam")
            except Exception:
                pass
            # 工业控制台配色：高对比、状态明确、少装饰。
            self.COL_ACCENT = "#0b5cab"
            self.COL_ACCENT_TXT = "#084b8a"
            self.COL_BG = "#eef2f5"
            self.COL_PANEL = "#ffffff"
            self.COL_BORDER = "#b8c4ce"
            self.COL_TEXT = "#17212b"
            self.COL_OK = "#16803c"
            self.COL_DANGER = "#c62828"
            self.root.configure(bg=self.COL_BG)
            style.configure(".", background=self.COL_BG, foreground=self.COL_TEXT,
                            font=(self.FONT, 10))
            style.configure("TLabelframe", background=self.COL_BG, bordercolor=self.COL_BORDER,
                            relief="groove")
            style.configure("TLabelframe.Label", background=self.COL_BG,
                            foreground=self.COL_ACCENT_TXT, font=(self.FONT, 10, "bold"))
            style.configure("TLabel", background=self.COL_BG, foreground=self.COL_TEXT)
            style.configure("TCheckbutton", background=self.COL_BG, foreground=self.COL_TEXT)
            style.map("TCheckbutton", background=[("active", self.COL_BG)])
            style.configure("TEntry", fieldbackground=self.COL_PANEL, padding=4)
            style.configure("TCombobox", fieldbackground=self.COL_PANEL, padding=3)
            style.configure("TSpinbox", fieldbackground=self.COL_PANEL, padding=3)
            style.configure("Accent.TButton", background=self.COL_ACCENT, foreground="white",
                            padding=(10, 5), font=(self.FONT, 10, "bold"))
            style.map("Accent.TButton",
                      background=[("active", "#084b8a"), ("disabled", "#9fb9d1")])
            style.configure("Test.TButton", background="#087f8c", foreground="white",
                            padding=(8, 5))
            style.map("Test.TButton", background=[("active", "#06636d"), ("disabled", "#9cc9ce")])
            style.configure("Danger.TButton", background=self.COL_DANGER, foreground="white",
                            padding=(10, 5), font=(self.FONT, 10, "bold"))
            style.map("Danger.TButton", background=[("active", "#9f1d1d"), ("disabled", "#d9a2a2")])
            style.configure("Treeview", rowheight=28, fieldbackground=self.COL_PANEL,
                            background=self.COL_PANEL, font=(self.FONT, 10))
            style.configure("Treeview.Heading", font=(self.FONT, 10, "bold"),
                            background="#d8e0e7", foreground="#102a43", padding=(6, 5))
            style.map("Treeview", background=[("selected", "#cfe3f5")],
                      foreground=[("selected", "#0b3157")])
        except Exception:
            pass

    def _build_ui(self):
        self._setup_style()
        # 顶部标题栏
        hdr = ttk.Frame(self.root)
        hdr.pack(fill="x", padx=12, pady=(10, 2))
        ttk.Label(hdr, text="工控机自动填数助手", font=(self.FONT, 16, "bold"),
                  foreground=self.COL_ACCENT_TXT).pack(side="left")
        ttk.Label(hdr, text="·  USB HID 注入（鼠标 + 键盘）", foreground="#64748b").pack(side="left", padx=(10, 0), pady=(4, 0))

        top = ttk.LabelFrame(self.root, text="连接与配置", padding=8)
        top.pack(fill="x", padx=8, pady=(6, 4))

        # 串口 / 波特率 / 测试
        ttk.Label(top, text="串口:").grid(row=0, column=0, sticky="e", padx=(4, 2))
        self.ui["port"] = ttk.Combobox(top, width=12, values=self._list_ports())
        self.ui["port"].grid(row=0, column=1, sticky="w")
        ttk.Button(top, text="刷新", width=6, command=self._refresh_ports).grid(row=0, column=2)
        ttk.Button(top, text="自动检测", width=8, command=self._auto_detect).grid(row=0, column=3)
        ttk.Label(top, text="波特率:").grid(row=0, column=4, sticky="e", padx=(16, 2))
        self.ui["baud"] = ttk.Combobox(top, width=8, values=["9600", "115200", "57600", "38400"])
        self.ui["baud"].set("115200")
        self.ui["baud"].grid(row=0, column=5, sticky="w")
        # 设备类型选择：ESP32-S3（当前固件）或 Pico (RP2040)。两者协议一致，仅作标识/持久化。
        ttk.Label(top, text="设备:").grid(row=0, column=6, sticky="e", padx=(10, 2))
        self.ui["device"] = ttk.Combobox(top, width=13, values=["ESP32-S3", "Pico (RP2040)"])
        self.ui["device"].set("ESP32-S3")
        self.ui["device"].grid(row=0, column=7, sticky="w")
        self.ui["btn_test"] = ttk.Button(top, text="测试连接", command=self._test_link)
        self.ui["btn_test"].grid(row=0, column=8, padx=8)
        self.ui["btn_hid"] = ttk.Button(top, text="HID 连接检测", command=self._check_hid)
        self.ui["btn_hid"].grid(row=0, column=9, padx=(0, 8))

        # 数据文件 / 工作表
        ttk.Label(top, text="数据文件:").grid(row=1, column=0, sticky="e", padx=(4, 2))
        self.ui["file"] = ttk.Entry(top, width=30)
        self.ui["file"].grid(row=1, column=1, columnspan=2, sticky="we")
        ttk.Button(top, text="浏览…", width=6, command=self._pick_file).grid(row=1, column=3, padx=(8, 0))
        ttk.Label(top, text="工作表:").grid(row=1, column=4, sticky="e", padx=(12, 2))
        self.ui["sheet"] = ttk.Combobox(top, width=12)
        self.ui["sheet"].grid(row=1, column=5, sticky="w")
        ttk.Button(top, text="载入表格", command=self._load_table).grid(row=1, column=6, padx=8)

        # 表头位置 / 是否含表头（用于读取非标准表格）
        ttk.Label(top, text="表头行:").grid(row=3, column=0, sticky="e", padx=(4, 2))
        self.ui["header_row"] = ttk.Spinbox(top, from_=0, to=1000, width=4)
        self.ui["header_row"].set(0)
        self.ui["header_row"].grid(row=3, column=1, sticky="w")
        self.ui["has_header"] = tk.BooleanVar(value=True)
        ttk.Checkbutton(top, text="有表头（第0行=第一行；去掉则按 col1/col2… 生成列名）",
                        variable=self.ui["has_header"]).grid(row=3, column=2, columnspan=3, sticky="w", padx=4)

        # 填充范围：要填的数据行
        ttk.Label(top, text="填充行:").grid(row=4, column=0, sticky="e", padx=(4, 2))
        self.ui["row_start"] = ttk.Spinbox(top, from_=1, to=999999, width=5)
        self.ui["row_start"].set(1)
        self.ui["row_start"].grid(row=4, column=1, sticky="w")
        ttk.Label(top, text="到第").grid(row=4, column=2, sticky="e")
        self.ui["row_end"] = ttk.Spinbox(top, from_=0, to=999999, width=5)
        self.ui["row_end"].set(0)
        self.ui["row_end"].grid(row=4, column=3, sticky="w")
        ttk.Label(top, text="行(0=全部)").grid(row=4, column=4, sticky="w", padx=(2, 0))
        self.ui["row_total"] = ttk.Label(top, text="尚未读取表格", foreground="#64748b")
        self.ui["row_total"].grid(row=4, column=5, columnspan=2, sticky="w", padx=(12, 0))

        # 坐标定位：先把光标移回左上角→归零，再用方向键挪动，对准后按 F8 写入所选步骤
        ttk.Label(top, text="坐标定位:").grid(row=5, column=0, sticky="e", padx=(4, 2))
        self.ui["btn_loc_home"] = ttk.Button(top, text="●归零", command=self._locate_home, width=6)
        self.ui["btn_loc_home"].grid(row=5, column=1, sticky="w")
        locnav = ttk.Frame(top)
        locnav.grid(row=5, column=2, columnspan=2, sticky="w")
        self.ui["btn_loc_up"] = ttk.Button(locnav, text="▲", width=3, command=lambda: self._locate_arrow(0, -1))
        self.ui["btn_loc_down"] = ttk.Button(locnav, text="▼", width=3, command=lambda: self._locate_arrow(0, 1))
        self.ui["btn_loc_left"] = ttk.Button(locnav, text="◀", width=3, command=lambda: self._locate_arrow(-1, 0))
        self.ui["btn_loc_right"] = ttk.Button(locnav, text="▶", width=3, command=lambda: self._locate_arrow(1, 0))
        # 按实体键盘方向键的 T 形布局：上在中间，左/下/右在下一排。
        self.ui["btn_loc_up"].grid(row=0, column=1, padx=1)
        self.ui["btn_loc_left"].grid(row=1, column=0, padx=1)
        self.ui["btn_loc_down"].grid(row=1, column=1, padx=1)
        self.ui["btn_loc_right"].grid(row=1, column=2, padx=1)
        ttk.Label(top, text="步长:").grid(row=5, column=4, sticky="e")
        self.ui["loc_step"] = ttk.Spinbox(top, from_=1, to=500, increment=10, width=5)
        self.ui["loc_step"].set(10)
        self.ui["loc_step"].grid(row=5, column=5, sticky="w")
        ttk.Button(top, text="F8 记录到所选步骤", command=self._locate_capture).grid(row=5, column=6, padx=8)
        self.ui["loc_pos"] = ttk.Label(top, text="当前光标: (—, —)", foreground="#64748b")
        self.ui["loc_pos"].grid(row=5, column=7, sticky="w", padx=(8, 0))
        # 直接输入目标坐标并跳转，便于快速定位，不必反复点击方向键。
        ttk.Label(top, text="跳转坐标:").grid(row=5, column=8, sticky="e", padx=(10, 2))
        self.ui["loc_target"] = ttk.Entry(top, width=13)
        self.ui["loc_target"].insert(0, "0,0")
        self.ui["loc_target"].grid(row=5, column=9, sticky="w")
        ttk.Button(top, text="跳转", width=6, command=self._locate_goto).grid(row=5, column=10, padx=(4, 0))
        self.root.bind_all("<F8>", lambda ev: self._locate_capture())

        # 选项
        self.ui["start_home"] = tk.BooleanVar(value=True)
        self.ui["confirm"] = tk.BooleanVar(value=True)
        # 流程运行前自动归零开关：置于坐标定位功能下方，与 start_home 配置项联动
        ttk.Checkbutton(top, text="每一步执行前自动归零（每步流程先回屏幕左上角原点，避免坐标漂移）",
                        variable=self.ui["start_home"]).grid(row=6, column=0, columnspan=9, sticky="w", padx=4, pady=(3, 0))
        ttk.Checkbutton(top, text="每行填完人工确认后再继续",
                        variable=self.ui["confirm"]).grid(row=2, column=3, columnspan=2, sticky="w")
        ttk.Label(top, text="步进延时(ms):").grid(row=2, column=5, sticky="e")
        self.ui["delay"] = ttk.Spinbox(top, from_=0, to=5000, increment=50, width=7)
        self.ui["delay"].set(300)
        self.ui["delay"].grid(row=2, column=6, sticky="w", padx=(4, 0))

        # ---- 无线控制（WiFi）：独立子容器，内部自带列宽，避免与主面板列冲突/被遮挡 ----
        # 背景：原先这些控件直接放在 row0/row2 —— row0 共 16 列总宽约 1700px 超出窗口可视宽度，
        # 且 row2 的 column3-6 已被「人工确认/步进延时」占用，导致同格重叠被盖住而看不见。
        # 排列紧凑（2 行）：窗口高度紧张，多一行都可能把底部日志区挤出可视范围。
        wf = ttk.Frame(top)
        wf.grid(row=7, column=0, columnspan=10, sticky="w", padx=4, pady=(8, 0))
        ttk.Label(wf, text="无线控制（WiFi）", font=(self.FONT, 10, "bold"),
                  foreground=self.COL_ACCENT_TXT).grid(row=0, column=0, sticky="w", padx=(0, 10))

        ttk.Label(wf, text="传输:").grid(row=0, column=1, sticky="e", padx=(0, 2))
        self.ui["transport"] = ttk.Combobox(wf, width=10, values=["串口", "无线(WiFi)"], state="readonly")
        self.ui["transport"].set("串口")
        self.ui["transport"].grid(row=0, column=2, sticky="w")
        ttk.Label(wf, text="IP:").grid(row=0, column=3, sticky="e", padx=(12, 2))
        self.ui["wifi_ip"] = ttk.Entry(wf, width=15)
        self.ui["wifi_ip"].grid(row=0, column=4, sticky="w")
        ttk.Label(wf, text="端口:").grid(row=0, column=5, sticky="e", padx=(12, 2))
        self.ui["wifi_port"] = ttk.Entry(wf, width=7)
        self.ui["wifi_port"].insert(0, "8080")
        self.ui["wifi_port"].grid(row=0, column=6, sticky="w")

        # WiFi 凭据（写入 ESP32 NVS，断电保存，无需改固件重烧）。先用串口连上板子再发。
        ttk.Label(wf, text="WiFi名(SSID):").grid(row=1, column=0, sticky="e", padx=(0, 2), pady=(4, 0))
        self.ui["wifi_ssid"] = ttk.Entry(wf, width=20)
        self.ui["wifi_ssid"].grid(row=1, column=1, columnspan=3, sticky="we", pady=(4, 0))
        ttk.Label(wf, text="密码:").grid(row=1, column=4, sticky="e", padx=(8, 2), pady=(4, 0))
        self.ui["wifi_pass"] = ttk.Entry(wf, width=18, show="*")
        self.ui["wifi_pass"].grid(row=1, column=5, columnspan=2, sticky="we", pady=(4, 0))
        self.ui["btn_wifi"] = ttk.Button(wf, text="保存并连接", command=self._save_wifi)
        self.ui["btn_wifi"].grid(row=1, column=7, padx=(10, 0), pady=(4, 0))
        ttk.Label(wf, text="先串口连板 → 写入NVS → 自动取IP并切到无线；笔记本需与 ESP32 同网段",
                  foreground="#64748b").grid(row=2, column=0, columnspan=8, sticky="w", pady=(2, 0))

        # 表格列选择 + 快速生成填入步骤
        colf = ttk.LabelFrame(self.root, text="选择要填入的表格列（可多选）", padding=8)
        colf.pack(fill="x", padx=8, pady=4)
        self.ui["col_list"] = tk.Listbox(colf, selectmode=tk.EXTENDED, height=4, font=(self.FONT, 10))
        self.ui["col_list"].grid(row=0, column=0, sticky="we", rowspan=2)
        cbs = ttk.Scrollbar(colf, orient="vertical", command=self.ui["col_list"].yview)
        cbs.grid(row=0, column=1, sticky="ns", rowspan=2)
        self.ui["col_list"].configure(yscrollcommand=cbs.set)
        ttk.Label(colf, text="在左侧按住 Ctrl/Shift 多选列，\n点下方按钮为每一列生成一个『输入数值』步骤。",
                  foreground="#64748b").grid(row=0, column=2, sticky="nw", padx=(10, 0))
        btns = ttk.Frame(colf)
        btns.grid(row=1, column=2, sticky="w", padx=(10, 0), pady=(6, 0))
        ttk.Button(btns, text="▶ 将选中列生成填入步骤", style="Accent.TButton",
                   command=self._gen_col_steps).pack(side="left")
        ttk.Button(btns, text="清空选择", command=self._clear_col_sel).pack(side="left", padx=(8, 0))
        colf.columnconfigure(0, weight=1)

        # 中间：步骤流程编辑器
        mid = ttk.LabelFrame(self.root, text="步骤流程（每行 = 一步；对每一行表格数据按顺序执行一遍）", padding=8)
        mid.pack(fill="both", expand=True, padx=8, pady=4)

        # 编辑对象选择 + 按钮管理
        objrow = ttk.Frame(mid)
        objrow.pack(fill="x", pady=(0, 4))
        ttk.Label(objrow, text="编辑对象:").pack(side="left")
        self.ui["obj"] = ttk.Combobox(objrow, width=20, state="readonly")
        self.ui["obj"].pack(side="left", padx=(4, 12))
        self.ui["obj"].bind("<<ComboboxSelected>>", self._on_obj_change)
        ttk.Button(objrow, text="＋新建按钮", command=self._add_button).pack(side="left", padx=2)
        ttk.Button(objrow, text="重命名按钮", command=self._rename_button).pack(side="left", padx=2)
        ttk.Button(objrow, text="删除按钮", command=self._delete_button).pack(side="left", padx=2)
        self.ui["obj_title"] = ttk.Label(objrow, text="主流程")
        self.ui["obj_title"].pack(side="right")

        # 步骤列表 + 动态表单
        panes = ttk.Frame(mid)
        panes.pack(fill="both", expand=True)

        treef = ttk.Frame(panes)
        treef.pack(side="left", fill="both", expand=True, padx=(0, 8))
        cols = ["type", "name", "detail"]
        self.ui["tree"] = ttk.Treeview(treef, columns=cols, show="headings", height=9)
        for c in cols:
            w = {"type": 100, "name": 150, "detail": 260}[c]
            self.ui["tree"].heading(c, text={"type": "类型", "name": "说明", "detail": "执行内容"}[c])
            self.ui["tree"].column(c, width=w, anchor="w")
        self.ui["tree"].pack(side="left", fill="both", expand=True)
        self.ui["tree"].bind("<<TreeviewSelect>>", self._on_tree_select)
        sb = ttk.Scrollbar(treef, orient="vertical", command=self.ui["tree"].yview)
        sb.pack(side="left", fill="y")
        self.ui["tree"].configure(yscrollcommand=sb.set)

        form = ttk.Frame(panes)
        form.pack(side="left", fill="y", padx=4)
        self._build_step_form(form)

        # 步骤编辑按钮
        btnrow = ttk.Frame(form)
        btnrow.grid(row=12, column=0, columnspan=2, sticky="we", pady=(4, 0))
        ttk.Button(btnrow, text="＋ 添加", command=self._add_step, width=8).pack(side="left", padx=2)
        ttk.Button(btnrow, text="更新所选", command=self._update_step, width=8).pack(side="left", padx=2)
        ttk.Button(btnrow, text="删除所选", command=self._del_step, width=8).pack(side="left", padx=2)
        ttk.Button(btnrow, text="↑上移", command=self._move_up, width=6).pack(side="left", padx=2)
        ttk.Button(btnrow, text="↓下移", command=self._move_down, width=6).pack(side="left", padx=2)
        ttk.Button(btnrow, text="▶ 测试所选步骤", command=self._test_selected_step,
                   style="Test.TButton", width=13).pack(side="left", padx=(10, 2))

        # 自定义按钮快捷执行区
        btnbar = ttk.LabelFrame(self.root, text="按钮快捷执行（点击即发送，可先归零）", padding=6)
        btnbar.pack(fill="x", padx=8, pady=4)
        self._btnbar_canvas = tk.Canvas(btnbar, height=48)
        self._btnbar_scroll = ttk.Scrollbar(btnbar, orient="horizontal", command=self._btnbar_canvas.xview)
        self._btnbar_canvas.configure(xscrollcommand=self._btnbar_scroll.set)
        self._btnbar_scroll.pack(fill="x")
        self._btnbar_canvas.pack(fill="x")
        self._btnbar_inner = ttk.Frame(self._btnbar_canvas)
        self._btnbar_canvas.create_window((0, 0), window=self._btnbar_inner, anchor="nw")
        self._btnbar_inner.bind("<Configure>",
                                lambda e: self._btnbar_canvas.configure(scrollregion=self._btnbar_canvas.bbox("all")))

        # 操作按钮
        ops = ttk.Frame(self.root)
        ops.pack(fill="x", padx=8, pady=4)
        self.ui["btn_save_cfg"] = ttk.Button(ops, text="保存配置", command=self._save_config)
        self.ui["btn_save_cfg"].pack(side="left")
        self.ui["btn_load_cfg"] = ttk.Button(ops, text="载入配置…", command=self._pick_config)
        self.ui["btn_load_cfg"].pack(side="left", padx=6)
        self.ui["btn_dry"] = ttk.Button(ops, text="Dry Run 预览", command=self._dry_run)
        self.ui["btn_dry"].pack(side="left", padx=(24, 0))
        self.ui["btn_cal"] = ttk.Button(ops, text="校准坐标", command=self._start_calibrate)
        self.ui["btn_cal"].pack(side="left", padx=6)
        self.ui["btn_start"] = ttk.Button(ops, text="▶ 开始填数 (F9)", style="Accent.TButton",
                                      command=self._start_fill)
        self.ui["btn_start"].pack(side="left", padx=(24, 0))
        self.ui["btn_stop"] = ttk.Button(ops, text="■ 停止 (F10)", command=self._stop_worker, style="Danger.TButton", state="disabled")
        self.ui["btn_stop"].pack(side="left", padx=6)
        self.ui["btn_confirm"] = ttk.Button(ops, text="✔ 已确认（空格/回车）",
                                            command=self._confirm_proceed, state="disabled")
        self.ui["btn_confirm"].pack(side="left", padx=(24, 0))

        # 进度
        prog = ttk.LabelFrame(self.root, text="进度", padding=6)
        prog.pack(fill="x", padx=8, pady=4)
        self.ui["row_label"] = ttk.Label(prog, text="尚未开始")
        self.ui["row_label"].pack(anchor="w")
        self.ui["bar"] = ttk.Progressbar(prog, mode="determinate")
        self.ui["bar"].pack(fill="x", pady=(4, 0))
        self.ui["summary"] = ttk.Label(prog, text="")
        self.ui["summary"].pack(anchor="w", pady=(4, 0))

        # 日志
        logf = ttk.LabelFrame(self.root, text="日志（Ctrl+C 可复制）", padding=6)
        logf.pack(fill="both", expand=True, padx=8, pady=(4, 8))
        logbar = ttk.Frame(logf)
        logbar.pack(fill="x", pady=(0, 2))
        ttk.Button(logbar, text="清空日志", width=9, command=self._clear_log).pack(side="left")
        ttk.Button(logbar, text="导出日志…", width=10, command=self._export_log).pack(side="left", padx=4)
        self.ui["log"] = scrolledtext.ScrolledText(logf, height=10, state="disabled",
                                                   font=(self.FONT, 10))
        self.ui["log"].pack(fill="both", expand=True)

    def _build_step_form(self, parent):
        """在 parent 里建步骤表单（2 列）。字段引用存进 self.ui。"""
        self.ui["f_type"] = ttk.Combobox(parent, width=12, values=[STEP_TYPE_LABELS[t] for t in STEP_TYPES],
                                         state="readonly")
        self.ui["f_type"].set(STEP_TYPE_LABELS["click"])
        self.ui["f_type"].bind("<<ComboboxSelected>>", lambda ev: self._refresh_step_form())
        self.ui["f_type"].bind("<<FocusOut>>", lambda ev: self._refresh_step_form())
        ttk.Label(parent, text="类型").grid(row=0, column=0, sticky="e", pady=2, padx=(0, 4))
        self.ui["f_type"].grid(row=0, column=1, sticky="w", pady=2)

        self.ui["f_name"] = self._form_entry(parent, 1, "说明")
        self.ui["f_x"] = self._form_entry(parent, 2, "X 坐标", width=8)
        self.ui["f_y"] = self._form_entry(parent, 2, "Y 坐标", width=8, second=True)
        self.ui["f_key"] = self._form_entry(parent, 3, "按键", width=8)
        self.ui["f_times"] = self._form_entry(parent, 3, "次数", width=8, second=True)
        # 表格列下拉：载入表格后自动填充候选项，也可手动输入列名或列序号
        self.ui["f_col"] = ttk.Combobox(parent, width=12)
        self.ui["f_col"].grid(row=4, column=1, sticky="w", pady=2)
        ttk.Label(parent, text="表格列").grid(row=4, column=0, sticky="e", pady=2, padx=(0, 4))
        self.ui["f_value"] = self._form_entry(parent, 4, "固定值", width=10, second=True)
        self.ui["f_pre"] = self._form_entry(parent, 5, "前置按键", width=12)
        self.ui["f_round"] = self._form_entry(parent, 5, "小数位", width=8, second=True)
        self.ui["f_ms"] = self._form_entry(parent, 6, "时长ms", width=8)
        self.ui["f_delta"] = self._form_entry(parent, 6, "滚轮", width=8, second=True)
        self.ui["f_delay"] = self._form_entry(parent, 7, "后延时", width=8)
        self.ui["f_enter"] = tk.BooleanVar(value=False)
        self.ui["f_enter_btn"] = ttk.Checkbutton(parent, text="输入后回车", variable=self.ui["f_enter"])
        self.ui["f_enter_btn"].grid(row=7, column=3, sticky="w")
        self.ui["f_prefix"] = self._form_entry(parent, 8, "前置文本", width=8)
        self.ui["f_suffix"] = self._form_entry(parent, 8, "后缀文本", width=8, second=True)
        self.ui["f_default"] = self._form_entry(parent, 9, "空值兜底", width=8)

        # 显示一个当前类型的字段说明
        self.ui["f_hint"] = ttk.Label(parent, text="", foreground="#666", wraplength=230)
        self.ui["f_hint"].grid(row=11, column=0, columnspan=4, sticky="w", pady=(4, 0))

        # 初始按默认类型（单击）置灰无关字段
        self._refresh_step_form()

    def _form_entry(self, parent, row, label, width=10, second=False):
        ent = ttk.Entry(parent, width=width)
        if second:
            ent.grid(row=row, column=3, sticky="w", pady=2)
            ttk.Label(parent, text=label).grid(row=row, column=2, sticky="e", pady=2, padx=(8, 4))
        else:
            ttk.Label(parent, text=label).grid(row=row, column=0, sticky="e", pady=2, padx=(0, 4))
            ent.grid(row=row, column=1, sticky="w", pady=2)
        return ent

    # ---------------- 步骤表单动态化 ----------------
    _FIELD_WIDGETS = {
        "name": "f_name", "x": "f_x", "y": "f_y", "key": "f_key", "times": "f_times",
        "col": "f_col", "value": "f_value", "pre": "f_pre", "round": "f_round",
        "ms": "f_ms", "delta": "f_delta", "delay": "f_delay",
        "prefix": "f_prefix", "suffix": "f_suffix", "default": "f_default",
    }

    def _refresh_step_form(self):
        label = self.ui["f_type"].get()
        typ = None
        for t, lab in STEP_TYPE_LABELS.items():
            if lab == label:
                typ = t
                break
        fields = FORM_FIELDS_BY_TYPE.get(typ, ())
        for key, wname in self._FIELD_WIDGETS.items():
            w = self.ui.get(wname)
            if w is None:
                continue
            w.configure(state="normal" if key in fields else "disabled")
        self.ui["f_enter_btn"].configure(state="normal" if "enter" in fields else "disabled")
        # hint
        hints = {
            "click": "移动光标到 (X,Y) 并单击：用于进入界面、点击按钮。",
            "dblclick": "移动到 (X,Y) 并双击。",
            "press_at": "移动到 (X,Y) 并按住左键「时长ms」毫秒（长按）。",
            "move": "只把光标移动到 (X,Y)，不点击。",
            "scroll": "滚动滚轮，「滚轮」正数向上滚、负数向下滚（如 -120）。",
            "key": "按一个键，如 enter/tab/esc/f5/ctrl+a/shift+tab；「次数」表示重复。",
            "input": "填入一个数值：优先取「表格列」这一行的值；没填列则用「固定值」。",
            "sleep": "原地等待「时长ms」毫秒。",
            "home": "让 Pico 把光标甩回屏幕左上角并重设原点（先归零再走坐标）。",
            "confirm": "到这里暂停，等人工点『已确认』再继续。",
            "button": "插入一个自定义按钮，执行时展开为按钮的步骤。",
        }
        self.ui["f_hint"].configure(text=hints.get(typ, ""))

    def _step_from_form(self):
        label = self.ui["f_type"].get()
        typ = next((t for t, lab in STEP_TYPE_LABELS.items() if lab == label), "click")
        s = {"type": typ, "name": self.ui["f_name"].get().strip()}
        try:
            if typ in MOUSE_STEP_TYPES:
                s["x"] = int(self.ui["f_x"].get() or 0)
                s["y"] = int(self.ui["f_y"].get() or 0)
                if typ == "press_at":
                    s["ms"] = int(self.ui["f_ms"].get() or 500)
            elif typ in ("scroll", "wheel"):
                s["delta"] = int(self.ui["f_delta"].get() or -120)
            elif typ == "key":
                s["key"] = self.ui["f_key"].get().strip()
                s["times"] = int(self.ui["f_times"].get() or 1)
            elif typ == "input":
                col = self.ui["f_col"].get().strip()
                val = self.ui["f_value"].get().strip()
                if col:
                    try:
                        s["col"] = int(col)     # 纯数字视为第 N 列（0-based）
                    except ValueError:
                        s["col"] = col          # 否则为列名
                if val:
                    s["value"] = val
                pre = [k.strip() for k in self.ui["f_pre"].get().split(",") if k.strip()]
                if pre:
                    s["pre_keys"] = pre
                r = self.ui["f_round"].get().strip()
                if r:
                    s["round"] = int(r)
                if self.ui["f_enter"].get():
                    s["enter"] = True
                pfx = self.ui["f_prefix"].get()
                if pfx:
                    s["prefix"] = pfx
                sfx = self.ui["f_suffix"].get()
                if sfx:
                    s["suffix"] = sfx
                dflt = self.ui["f_default"].get()
                if dflt:
                    s["default"] = dflt
            elif typ == "sleep":
                s["ms"] = int(self.ui["f_ms"].get() or 0)
            if "delay" in FORM_FIELDS_BY_TYPE.get(typ, ()):
                dm = self.ui["f_delay"].get().strip()
                if dm:
                    s["delay_ms"] = int(dm)
        except ValueError:
            messagebox.showerror("格式错误", "X/Y 坐标、次数、小数位、时长、后延时、滚轮必须是整数。")
            return None
        return s

    def _fill_form_from_step(self, s):
        typ = s.get("type", "click")
        self.ui["f_type"].set(STEP_TYPE_LABELS.get(typ, typ))
        self._refresh_step_form()
        self.ui["f_name"].delete(0, "end"); self.ui["f_name"].insert(0, s.get("name", ""))
        self.ui["f_x"].delete(0, "end"); self.ui["f_x"].insert(0, str(s.get("x", "")))
        self.ui["f_y"].delete(0, "end"); self.ui["f_y"].insert(0, str(s.get("y", "")))
        self.ui["f_key"].delete(0, "end"); self.ui["f_key"].insert(0, s.get("key", ""))
        self.ui["f_times"].delete(0, "end"); self.ui["f_times"].insert(0, str(s.get("times", "1")))
        self.ui["f_col"].set(s.get("col", ""))
        self.ui["f_value"].delete(0, "end"); self.ui["f_value"].insert(0, s.get("value", ""))
        self.ui["f_pre"].delete(0, "end"); self.ui["f_pre"].insert(0, ",".join(s.get("pre_keys", [])))
        self.ui["f_round"].delete(0, "end"); self.ui["f_round"].insert(0, "" if s.get("round") is None else str(s["round"]))
        self.ui["f_ms"].delete(0, "end"); self.ui["f_ms"].insert(0, str(s.get("ms", "")))
        self.ui["f_delta"].delete(0, "end"); self.ui["f_delta"].insert(0, str(s.get("delta", "-120")))
        self.ui["f_delay"].delete(0, "end"); self.ui["f_delay"].insert(0, str(s.get("delay_ms", "")))
        self.ui["f_prefix"].delete(0, "end"); self.ui["f_prefix"].insert(0, s.get("prefix", ""))
        self.ui["f_suffix"].delete(0, "end"); self.ui["f_suffix"].insert(0, s.get("suffix", ""))
        self.ui["f_default"].delete(0, "end"); self.ui["f_default"].insert(0, s.get("default", ""))
        self.ui["f_enter"].set(bool(s.get("enter")))

    # ---------------- 步骤列表维护 ----------------
    def _refresh_tree(self):
        t = self.ui["tree"]
        t.delete(*t.get_children())
        for idx, s in enumerate(self.objects.get(self.current_obj) or []):
            t.insert("", "end", iid=str(idx), values=[STEP_TYPE_LABELS.get(s.get("type"), s.get("type")),
                                                     s.get("name", ""), _step_detail(s)])

    def _selected_index(self):
        sel = self.ui["tree"].selection()
        if not sel:
            return None
        try:
            return int(sel[0])
        except ValueError:
            return None

    def _refresh_current_title(self):
        self.ui["obj_title"].configure(text=self.object_labels.get(self.current_obj, ""))

    def _add_step(self):
        s = self._step_from_form()
        if s is None:
            return
        self.objects.setdefault(self.current_obj, []).append(s)
        self._mark_dirty()
        self._refresh_tree()
        self.ui["tree"].selection_set(str(len(self.objects[self.current_obj]) - 1))

    def _update_step(self):
        idx = self._selected_index()
        if idx is None:
            messagebox.showinfo("提示", "请先在列表中选中一行。")
            return
        s = self._step_from_form()
        if s is None:
            return
        self.objects[self.current_obj][idx] = s
        self._mark_dirty()
        self._refresh_tree()
        self.ui["tree"].selection_set(str(idx))

    def _del_step(self):
        objlist = self.objects.get(self.current_obj, [])
        sel = self.ui["tree"].selection()
        for iid in reversed([int(i) for i in sel]):
            if 0 <= iid < len(objlist):
                objlist.pop(iid)
        self._mark_dirty()
        self._refresh_tree()

    def _move_step(self, delta):
        idx = self._selected_index()
        if idx is None:
            return
        objlist = self.objects[self.current_obj]
        new = idx + delta
        if new < 0 or new >= len(objlist):
            return
        objlist[idx], objlist[new] = objlist[new], objlist[idx]
        self._mark_dirty()
        self._refresh_tree()
        self.ui["tree"].selection_set(str(new))

    def _move_up(self):
        self._move_step(-1)

    def _move_down(self):
        self._move_step(1)

    def _on_tree_select(self, _ev=None):
        idx = self._selected_index()
        if idx is None:
            return
        objlist = self.objects.get(self.current_obj, [])
        if 0 <= idx < len(objlist):
            self._fill_form_from_step(objlist[idx])

    # ---------------- 编辑对象（主流程 / 自定义按钮） ----------------
    def _on_obj_change(self, _ev=None):
        label = self.ui["obj"].get()
        for key, lab in self.object_labels.items():
            if lab == label:
                self.current_obj = key
                break
        self._refresh_tree()
        self._refresh_current_title()

    def _sync_obj_combo(self):
        self.ui["obj"]["values"] = list(self.object_labels.values())
        self.ui["obj"].set(self.object_labels.get(self.current_obj, "主流程"))

    def _add_button(self):
        name = simpledialog.askstring("新建按钮", "输入按钮名称（例如：进入压力界面）", parent=self.root)
        if not name:
            return
        name = name.strip()
        key = "btn:" + name
        if key in self.objects:
            messagebox.showwarning("提示", "已存在同名按钮：%s" % name)
            return
        self.objects[key] = []
        self.object_labels[key] = "按钮：" + name
        self._mark_dirty()
        self.current_obj = key
        self._sync_obj_combo()
        self._refresh_tree()
        self._refresh_current_title()
        self._rebuild_button_bar()
        self.log("已新建按钮：%s（请在步骤列表里添加它的步骤）" % name)

    def _rename_button(self):
        if not self.current_obj.startswith("btn:"):
            messagebox.showinfo("提示", "请先在『编辑对象』里选中一个按钮。")
            return
        old_name = self.current_obj[4:]
        new_name = simpledialog.askstring("重命名按钮", "新的按钮名称：", initialvalue=old_name, parent=self.root)
        if not new_name:
            return
        new_name = new_name.strip()
        new_key = "btn:" + new_name
        if new_key in self.objects and new_key != self.current_obj:
            messagebox.showwarning("提示", "已存在同名按钮：%s" % new_name)
            return
        self.objects[new_key] = self.objects.pop(self.current_obj)
        del self.object_labels[self.current_obj]
        self.object_labels[new_key] = "按钮：" + new_name
        self._mark_dirty()
        self.current_obj = new_key
        self._sync_obj_combo()
        self._refresh_current_title()
        self._rebuild_button_bar()
        self.log("按钮已重命名为：%s" % new_name)

    def _delete_button(self):
        if not self.current_obj.startswith("btn:"):
            messagebox.showinfo("提示", "请先在『编辑对象』里选中一个按钮。")
            return
        name = self.current_obj[4:]
        if not messagebox.askyesno("删除按钮", "确定删除按钮：%s？" % name):
            return
        del self.objects[self.current_obj]
        del self.object_labels[self.current_obj]
        self._mark_dirty()
        self.current_obj = "__flow__"
        self._sync_obj_combo()
        self._refresh_tree()
        self._refresh_current_title()
        self._rebuild_button_bar()
        self.log("已删除按钮：%s" % name)

    def _rebuild_button_bar(self):
        for child in self._btnbar_inner.winfo_children():
            child.destroy()
        btn_keys = [k for k in self.objects if k.startswith("btn:")]
        # 按加入顺序重建
        for key in btn_keys:
            name = self.object_labels[key][3:]  # 去掉 "按钮：" 前缀
            b = ttk.Button(self._btnbar_inner, text=name, command=lambda n=name: self._run_button(n))
            b.pack(side="left", padx=3, pady=3)

    # ---------------- 配置与表格 ----------------
    def _list_ports(self):
        try:
            from serial.tools import list_ports
            return [p.device for p in list_ports.comports()]
        except Exception:
            return []

    def _refresh_ports(self):
        self.ui["port"]["values"] = self._list_ports()

    def _auto_detect(self):
        if self.running:
            messagebox.showinfo("提示", "任务运行中，请先『停止』再检测。")
            return
        self._loc_close_link()  # 释放定位连接，避免端口冲突
        self.log("正在自动检测设备串口…（逐个端口 ping，可能需要几秒）")
        threading.Thread(target=self._detect_worker, daemon=True).start()

    def _detect_worker(self):
        ports = self._list_ports()
        if not ports:
            self.log("未发现任何串口，请检查 USB 连接。")
            return
        try:
            baud = int(self.ui["baud"].get())
        except ValueError:
            baud = 115200
        for p in ports:
            self.log("  尝试 %s …" % p)
            try:
                link = _open_link_bounded({"com_port": p, "baudrate": baud})
            except Exception:
                self.log("    %s 打开失败/超时" % p)
                continue
            self.current_link = link
            try:
                ok = link.ping()
            finally:
                if self.current_link is link:
                    self.current_link = None
                link.close()
            if ok:
                self.log("✅ 检测到 Pico：%s" % p)
                self.queue.put(("ui", ("port_set", p)))
                return
        self.log("未检测到 Pico：请检查接线（TX/RX）与驱动，或手动选择串口。")

    def _gen_col_steps(self):
        """把左侧勾选的表格列，逐个生成『输入数值』步骤，追加到当前流程对象。"""
        sel = self.ui["col_list"].curselection()
        if not sel:
            messagebox.showinfo("提示", "请先在左侧列表中选中要填入的表格列（可 Ctrl/Shift 多选）。")
            return
        names = [self.ui["col_list"].get(i).strip() for i in sel]
        names = [n for n in names if n and not n.startswith("（")]
        if not names:
            messagebox.showinfo("提示", "请先载入表格，再选择列。")
            return
        obj = self.current_obj
        steps = self.objects.setdefault(obj, [])
        for name in names:
            steps.append({"type": "input", "name": "填入" + name, "col": name,
                          "pre_keys": ["ctrl+a"], "enter": True})
        self._refresh_tree()
        self.log("已为 %d 列生成填入步骤并追加到【%s】：%s"
                 % (len(names), self.object_labels.get(obj, ""), "、".join(names)))
        self.ui["tree"].selection_set(str(len(steps) - 1))
        self._mark_dirty()

    def _clear_col_sel(self):
        self.ui["col_list"].selection_clear(0, "end")

    def _set_col_list(self, columns):
        """把表头列名填进左侧多选列表；无数据时给出提示。"""
        lb = self.ui["col_list"]
        lb.delete(0, "end")
        if columns:
            for c in columns:
                lb.insert("end", c)
        else:
            lb.insert("end", "（载入表格后此处列出所有列名）")

    def _pick_file(self):
        p = filedialog.askopenfilename(
            initialdir=HERE, title="选择数据文件",
            filetypes=[("表格", "*.xlsx *.xlsm *.xls *.csv"), ("所有文件", "*.*")])
        if p:
            self.ui["file"].delete(0, "end")
            self.ui["file"].insert(0, p)
            self._populate_sheets(p)
            self.ui["f_col"]["values"] = ()
            self._col_choices = []

    def _populate_sheets(self, path=None):
        """从工作簿里列出所有 sheet，填到工作表下拉。"""
        path = path or self.ui["file"].get().strip()
        if not path:
            return
        ext = os.path.splitext(path)[1].lower()
        try:
            if ext in (".xlsx", ".xlsm"):
                from openpyxl import load_workbook
                wb = load_workbook(path, read_only=True)
                sheets = list(wb.sheetnames)
                wb.close()
            elif ext == ".xls":
                import xlrd
                sheets = xlrd.open_workbook(path).sheet_names()
            else:
                return
        except Exception as e:
            self.log("读取工作表列表失败：%s" % e)
            return
        if sheets:
            self.ui["sheet"]["values"] = sheets
            if not self.ui["sheet"].get():
                self.ui["sheet"].set(sheets[0])

    def _int_field(self, wname, label, default=0):
        """安全读取整数 Spinbox/Entry：空值用默认，非法值抛中文提示。"""
        try:
            return int(self.ui[wname].get().strip() or default)
        except ValueError:
            raise ValueError("%s 必须是整数（当前值：%r）" % (label, self.ui[wname].get()))

    def collect_config(self):
        steps = copy.deepcopy(self.objects.get("__flow__", []))
        buttons = {}
        for key, ksteps in self.objects.items():
            if key.startswith("btn:"):
                buttons[key[4:]] = copy.deepcopy(ksteps)
        return {
            "com_port": self.ui["port"].get().strip(),
            "baudrate": self._int_field("baud", "波特率", 115200),
            "transport": "wifi" if self.ui["transport"].get().startswith("无线") else "serial",
            "wifi_ip": self.ui["wifi_ip"].get().strip(),
            "wifi_port": self._int_field("wifi_port", "WiFi 端口", 8080),
            "data_file": self.ui["file"].get().strip(),
            "sheet": self.ui["sheet"].get().strip() or None,
            "header_row": self._int_field("header_row", "表头行", 0),
            "has_header": self.ui["has_header"].get(),
            "row_start": self._int_field("row_start", "起始行", 1),
            "row_end": self._int_field("row_end", "结束行", 0),
            "confirm": "enter" if self.ui["confirm"].get() else "none",
            "start_home": self.ui["start_home"].get(),
            "step_delay_ms": self._int_field("delay", "步进延时(ms)", 0),
            "steps": steps,
            "buttons": buttons,
            "device_type": self._device_key(),
        }

    def _device_key(self):
        """把下拉框文本归一化为配置键：esp32 / pico。"""
        v = (self.ui["device"].get() if "device" in self.ui else "ESP32-S3").lower()
        return "pico" if "pico" in v else "esp32"

    def _has_link(self, cfg):
        """是否配置了可用链路：串口需 com_port，无线需 wifi_ip。"""
        if (cfg.get("transport") or "serial") == "wifi":
            return bool(cfg.get("wifi_ip"))
        return bool(cfg.get("com_port"))

    def _link_cfg(self, cfg):
        """抽取连接相关字段，供 worker cfg 构造复用（保证无线模式也能拿对链路）。"""
        return {
            "transport": cfg.get("transport", "serial"),
            "com_port": cfg.get("com_port", ""),
            "baudrate": cfg.get("baudrate", 115200),
            "wifi_ip": cfg.get("wifi_ip", ""),
            "wifi_port": cfg.get("wifi_port", 8080),
        }

    def _link_desc(self, cfg):
        """链路描述文本，用于日志（串口显示 COM 口，无线显示 IP:端口）。"""
        if (cfg.get("transport") or "serial") == "wifi":
            return "%s:%s" % (cfg.get("wifi_ip"), cfg.get("wifi_port", 8080))
        return cfg.get("com_port", "")

    def _toast(self, title, msg, kind="info"):
        """关键结果弹窗（线程安全）。

        窗口高度不足时 Tk 会直接『不映射』放不下的底部区块，日志区可能整个不可见，
        只写日志会让用户以为没反应。所以连接/配网这类关键结果必须同时弹窗反馈。
        """
        def _show():
            try:
                {"info": messagebox.showinfo, "warn": messagebox.showwarning,
                 "error": messagebox.showerror}.get(kind, messagebox.showinfo)(title, msg)
            except Exception:
                pass
        try:
            self.root.after(0, _show)
        except Exception:
            pass

    def apply_cfg_to_ui(self, cfg):
        self.ui["port"].set(cfg.get("com_port", ""))
        self.ui["baud"].set(str(cfg.get("baudrate", 115200)))
        self.ui["transport"].set("无线(WiFi)" if cfg.get("transport") == "wifi" else "串口")
        self.ui["wifi_ip"].delete(0, "end"); self.ui["wifi_ip"].insert(0, cfg.get("wifi_ip", ""))
        self.ui["wifi_port"].delete(0, "end"); self.ui["wifi_port"].insert(0, str(cfg.get("wifi_port", 8080)))
        self.ui["file"].delete(0, "end"); self.ui["file"].insert(0, cfg.get("data_file", ""))
        self.ui["sheet"].set(cfg.get("sheet") or "")
        self.ui["header_row"].set(cfg.get("header_row", 0))
        self.ui["has_header"].set(cfg.get("has_header", True))
        self.ui["row_start"].set(cfg.get("row_start", 1))
        self.ui["row_end"].set(cfg.get("row_end", 0))
        self.ui["confirm"].set(cfg.get("confirm", "enter") == "enter")
        self.ui["start_home"].set(cfg.get("start_home", True))
        self.ui["delay"].set(cfg.get("step_delay_ms", 300))
        self.ui["device"].set("Pico (RP2040)" if cfg.get("device_type") == "pico" else "ESP32-S3")
        self.device_type = cfg.get("device_type", "esp32")
        # 重建对象
        flow = cfg.get("steps")
        if not flow and cfg.get("targets"):
            flow = run._steps_from_targets(cfg["targets"])
        self.objects = {"__flow__": copy.deepcopy(flow or [])}
        self.object_labels = {"__flow__": "主流程"}
        for bname, bsteps in (cfg.get("buttons") or {}).items():
            self.objects["btn:" + bname] = copy.deepcopy(bsteps)
            self.object_labels["btn:" + bname] = "按钮：" + bname
        self.current_obj = "__flow__"
        self._sync_obj_combo()
        self._refresh_tree()
        self._refresh_current_title()
        self._rebuild_button_bar()
        self._mark_clean()

    def _load_default_config(self):
        if run is None:
            return
        if os.path.exists(self.cfg_path):
            try:
                self.apply_cfg_to_ui(run.read_config(self.cfg_path))
                return
            except Exception:
                pass
        example = os.path.join(HERE, "config.example.json")
        if os.path.exists(example):
            self.apply_cfg_to_ui(run.read_config(example))

    def _write_config(self, cfg, path=None):
        """原子写配置：临时文件 + os.replace，写前自动备份为 .bak。"""
        path = path or self.cfg_path
        tmp = path + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
        if os.path.exists(path):
            try:
                shutil.copy2(path, path + ".bak")
            except Exception:
                pass
        os.replace(tmp, path)

    def _save_config(self):
        try:
            cfg = self.collect_config()
            self._write_config(cfg)
            self._mark_clean()
            self.log("配置已保存到 %s" % self.cfg_path)
        except Exception as e:
            messagebox.showerror("保存失败", str(e))

    def _pick_config(self):
        p = filedialog.askopenfilename(initialdir=HERE, title="选择配置文件",
                                       filetypes=[("JSON", "*.json")])
        if not p:
            return
        try:
            self.apply_cfg_to_ui(run.read_config(p))
            self.cfg_path = p
            self.log("已载入配置: %s" % p)
        except Exception as e:
            messagebox.showerror("载入失败", str(e))

    def _load_table(self):
        cfg = self.collect_config()
        try:
            self._populate_sheets()
            header, records = run.load_records(cfg)
        except Exception as e:
            messagebox.showerror("读取表格失败", str(e))
            return
        self._col_choices = list(header)
        self.ui["f_col"]["values"] = header
        self._set_col_list(header)
        n = len(records)
        self.ui["row_total"].configure(text="已读取 %d 行数据" % n)
        self.log("表格读取成功：表头 %s，数据 %d 行" % (", ".join(header), n))
        # 校验步骤引用的列名（跳过整数列序号）
        bad = []
        for s in run.expand_steps(cfg):
            if s.get("type") == "input" and s.get("col") and not isinstance(s.get("col"), int):
                if s.get("col") not in header:
                    bad.append(s.get("col"))
        if bad:
            messagebox.showwarning("列不存在", "步骤引用的表格列不在表头中：%s。请检查表格或步骤设置。"
                                   % ", ".join(sorted(set(bad))))

    # ---------------- 状态标记 / 通用小工具 ----------------
    def _mark_dirty(self):
        """标记有未保存的修改，并在窗口标题上打星号。"""
        self.dirty = True
        self.root.title("工控机自动填数助手 *")

    def _mark_clean(self):
        self.dirty = False
        self.root.title("工控机自动填数助手")

    def _beep(self, count=1):
        try:
            import winsound
            for i in range(count):
                winsound.Beep(880 if i == 0 else 660, 150)
                if i < count - 1:
                    time.sleep(0.05)
        except Exception:
            pass

    def _cfg_or_error(self):
        """安全收集配置：collect_config 抛错（如波特率非数字）时弹窗并返回 None。"""
        try:
            return self.collect_config()
        except ValueError as e:
            messagebox.showerror("配置错误", str(e))
            return None

    # ---------------- 日志 / 消息泵 ----------------
    def log(self, msg):
        self.queue.put(("log", "%s %s" % (time.strftime("%H:%M:%S"), msg)))

    def _append_log(self, msg):
        w = self.ui["log"]
        w.configure(state="normal")
        w.insert("end", msg + "\n")
        # 防止长时间运行后日志无限膨胀：超过 3000 行时裁掉头部
        try:
            n = int(w.index("end-1c").split(".")[0])
            if n > 3000:
                w.delete("1.0", "%d.0" % (n - 2000))
        except Exception:
            pass
        w.see("end")
        w.configure(state="disabled")

    def _clear_log(self):
        w = self.ui["log"]
        w.configure(state="normal")
        w.delete("1.0", "end")
        w.configure(state="disabled")

    def _export_log(self):
        p = filedialog.asksaveasfilename(initialdir=HERE, title="导出日志",
                                         defaultextension=".log",
                                         filetypes=[("日志", "*.log"), ("文本", "*.txt")])
        if not p:
            return
        try:
            content = self.ui["log"].get("1.0", "end-1c")
            with open(p, "w", encoding="utf-8") as f:
                f.write(content)
            self.log("日志已导出：%s" % p)
        except Exception as e:
            messagebox.showerror("导出失败", str(e))

    def _poll_queue(self):
        try:
            while True:
                kind, payload = self.queue.get_nowait()
                if kind == "log":
                    self._append_log(payload)
                elif kind == "interact":
                    self._handle_interact(payload)
                elif kind == "ui":
                    self._apply_ui(payload)
        except queue.Empty:
            pass
        self.root.after(100, self._poll_queue)

    def _apply_ui(self, payload):
        key, value = payload
        try:
            if key == "running":
                self.running = value
                self.ui["btn_start"].configure(state="disabled" if value else "normal")
                self.ui["btn_stop"].configure(state="normal" if value else "disabled")
                self.ui["btn_cal"].configure(state="disabled" if value else "normal")
                self.ui["btn_dry"].configure(state="disabled" if value else "normal")
                self.ui["btn_test"].configure(state="disabled" if value else "normal")
                self.ui["btn_hid"].configure(state="disabled" if value else "normal")
                self.ui["btn_confirm"].configure(state="disabled")
                if not value and self.stop_event.is_set():
                    self.ui["row_label"].configure(text="已停止")
                    self.ui["summary"].configure(text="")
            elif key == "confirm_wait":
                self.ui["btn_confirm"].configure(state="normal" if value else "disabled")
            elif key == "row_label":
                self.ui["row_label"].configure(text=value)
            elif key == "summary":
                self.ui["summary"].configure(text=value)
            elif key == "bar_max":
                self.ui["bar"].configure(maximum=value)
            elif key == "bar":
                self.ui["bar"].configure(value=value)
            elif key == "tree_reload":
                self.apply_cfg_to_ui(value)
            elif key == "port_set":
                self.ui["port"].set(value)
                self.log("已自动选择串口: %s" % value)
            elif key == "beep":
                self._beep(int(value or 1))
            elif key == "beep_done":
                self._beep(3)
        except Exception:
            pass

    def _handle_interact(self, req):
        kind = req["kind"]
        if kind == "ask_yesno":
            val = messagebox.askyesno(req["title"], req["prompt"])
        elif kind == "ask_string":
            val = simpledialog.askstring(req["title"], req["prompt"], initialvalue=req.get("init", ""),
                                         parent=self.root)
        else:
            val = None
        if self.interact is None:
            return
        self.interact.resolve(req, val)

    # ---------------- 串口测试 ----------------
    def _test_link(self):
        self._loc_close_link()  # 释放定位连接，避免端口被占用导致测试打不开
        cfg = self._cfg_or_error()
        if cfg is None:
            return
        self.log("正在测试 %s 连接（%s）…" % (self._device_key().upper(), self._link_desc(cfg)))
        thr = threading.Thread(target=self._test_worker, args=(cfg,), daemon=True)
        thr.start()

    def _test_worker(self, cfg):
        dev = self._device_key().upper()
        link = None
        try:
            link = _open_link_bounded(cfg)
            self.current_link = link
            ok = link.ping()
            self.log("%s 在线 ✅（%s）" % (dev, self._link_desc(cfg)) if ok
                     else "%s 无响应 ❌（检查链路、接线或固件）" % dev)
            if ok:
                self._toast("连接测试", "%s 在线 ✅\n链路：%s" % (dev, self._link_desc(cfg)))
            else:
                self._toast("连接测试", "%s 无响应 ❌\n链路：%s\n\n检查链路、接线或固件。" %
                            (dev, self._link_desc(cfg)), kind="error")
        except Exception as e:
            self.log("连接失败: %s" % e)
            self._toast("连接测试", "连接失败：\n%s" % e, kind="error")
        finally:
            if self.current_link is link:
                self.current_link = None
            if link:
                link.close()

    # ---------------- HID 连接检测（原生 USB → 工控机） ----------------
    def _check_hid(self):
        self._loc_close_link()  # 释放定位连接，避免端口被占用
        cfg = self._cfg_or_error()
        if cfg is None:
            return
        thr = threading.Thread(target=self._check_hid_worker, args=(cfg,), daemon=True)
        thr.start()

    def _check_hid_worker(self, cfg):
        dev = self._device_key().upper()
        link = None
        try:
            link = _open_link_bounded(cfg)
            self.current_link = link
            if not link.ping():
                self.log("%s 链路无响应 ❌（先解决连接/固件，再测 HID）" % dev)
                return
            resp = link.send({"op": "usb_state"})
            mounted = bool(resp.get("mounted"))
            self._toast("HID 连接检测",
                        ("✅ 工控机已枚举到 %s 的原生 USB 键鼠复合设备。\n\n"
                         "若工控机仍『无反应』，问题在工控机侧：屏幕是否锁定、光标是否被甩到别的显示器、"
                         "目标窗口是否激活。" % dev) if mounted else
                        ("⚠️ 工控机未枚举到 %s 的 USB 键鼠设备。\n\n"
                         "OTG→工控机 物理链路没通，依次排查：\n"
                         "① OTG 线是否只充电无数据（最常见）→ 换带数据线的 USB 线；\n"
                         "② 工控机 USB 口是否损坏/禁用 → 换主板背面口；\n"
                         "③ 重插拔 OTG 两端。" % dev),
                        kind="info" if mounted else "warn")
            if mounted:
                self.log("✅ HID 已就绪：工控机已枚举到 %s 的原生 USB 键鼠复合设备。" % dev)
                self.log("   既然 HID 已就绪但工控机仍『无反应』，问题在工控机侧消费：")
                self.log("   ① 工控机屏幕是否锁定（登录界面不接收应用层输入）；")
                self.log("   ② 多显示器时光标被甩到了别的屏幕；③ 目标软件窗口未处于激活/最前。")
            else:
                self.log("⚠️ HID 未就绪：工控机没有枚举到 %s 的原生 USB 键鼠设备。" % dev)
                self.log("   说明 OTG→工控机 这条物理链路没通，按以下顺序排查：")
                self.log("   ① OTG 线是否只充电无数据——换一根确认带数据线的 USB 线（最常见）；")
                self.log("   ② 工控机 USB 口是否损坏/被禁用——换插机箱背面主板口，或在设备管理器看有无『未知 USB 设备』；")
                self.log("   ③ 接触不良——重插拔 OTG 口两端；")
                self.log("   ④ 若把 OTG 临时改插『本机』，本机光标能动/能打字，则证明固件与线材都正常，问题在工控机口。")
        except Exception as e:
            self.log("HID 检测失败: %s" % e)
        finally:
            if self.current_link is link:
                self.current_link = None
            if link:
                link.close()

    # ---------------- WiFi 凭据（写入 ESP32 NVS，断电保存） ----------------
    def _save_wifi(self):
        ssid = self.ui["wifi_ssid"].get().strip()
        passwd = self.ui["wifi_pass"].get()
        if not ssid:
            messagebox.showinfo("提示", "请先填写 WiFi 名称(SSID)。")
            return
        cfg = self._cfg_or_error()
        if cfg is None:
            return
        # 保存 WiFi 必须先经串口连上板子（板子尚未联网时只能走 UART/COM）。
        if not cfg.get("com_port"):
            messagebox.showinfo("提示",
                "保存 WiFi 需要先『串口』连接 ESP32（板子还没联网，只能经 COM 下发）。\n"
                "请先在『串口』处选好 COM 口再点保存并连接。")
            return
        serial_cfg = dict(cfg)
        serial_cfg["transport"] = "serial"   # 强制走串口下发 set_wifi
        thr = threading.Thread(target=self._save_wifi_worker, args=(serial_cfg, ssid, passwd), daemon=True)
        thr.start()

    def _save_wifi_worker(self, cfg, ssid, passwd):
        dev = self._device_key().upper()
        link = None
        try:
            link = _open_link_bounded(cfg)
            self.current_link = link
            if not link.ping():
                self.log("%s 串口无响应 ❌，无法下发 WiFi 配置（确认 COM 口与板子已上电）。" % dev)
                self._toast("配网失败", "%s 串口无响应 ❌，无法下发 WiFi 配置。\n\n"
                            "确认 COM 口选对、板子已上电，且操作台没被别的程序占用该端口。" % dev,
                            kind="error")
                return
            resp = link.send({"op": "set_wifi", "ssid": ssid, "pass": passwd})
            if resp.get("ack") != "ok":
                self.log("❌ WiFi 配置下发失败: %s" % resp.get("msg", resp))
                self._toast("配网失败", "WiFi 配置下发失败：\n%s" % resp.get("msg", resp), kind="error")
                return
            self.log("✅ WiFi 凭据已写入 %s 的 NVS（断电保存），正在联网…" % dev)
            # 轮询 net_info，最多 ~20s 等待拿到 IP
            ip = None
            for i in range(10):
                try:
                    r = link.send({"op": "net_info"}, timeout=3.0)
                except Exception:
                    r = {}
                if r.get("wifi") is True:
                    ip = r.get("ip")
                    break
                time.sleep(2.0)
            if ip:
                self.root.after(0, lambda: self.ui["wifi_ip"].delete(0, "end"))
                self.root.after(0, lambda: self.ui["wifi_ip"].insert(0, ip))
                self.root.after(0, lambda: self.ui["transport"].set("无线(WiFi)"))
                self.log("🌐 已连上 WiFi，ESP32 IP = %s（已自动填入『IP』并把传输切到无线）。" % ip)
                self.log("   现在点『测试连接』即可经 WiFi 控制；串口仍作兜底保留。")
                self._toast("配网成功",
                            "已连上 WiFi ✅\nESP32 IP = %s\n\n已自动填入『IP』并把传输切到『无线(WiFi)』；\n"
                            "点『测试连接』即可经 WiFi 控制，串口仍作兜底保留。" % ip)
            else:
                self.log("⚠️ 凭据已保存，但 %s 在 20s 内未连上 WiFi。" % dev)
                self.log("   可能原因：SSID/密码错、信号弱、或该网络禁止新设备。请核对后重试『保存并连接』。")
                self._toast("未连上 WiFi",
                            "凭据已保存，但 %s 在 20 秒内未连上 WiFi。\n\n"
                            "可能原因：SSID/密码错、信号弱、或该网络禁止新设备。\n"
                            "请核对后重试『保存并连接』。" % dev, kind="warn")
        except Exception as e:
            self.log("WiFi 保存失败: %s" % e)
            self._toast("配网失败", "WiFi 保存失败：\n%s" % e, kind="error")
        finally:
            if self.current_link is link:
                self.current_link = None
            if link:
                link.close()

    # ---------------- 坐标定位（F8 记录） ----------------
    def _locate_get_cfg(self):
        if self.running:
            messagebox.showinfo("提示", "任务运行中，请先『停止』再定位坐标。")
            return None
        cfg = self._cfg_or_error()
        if cfg is None:
            return None
        if not self._has_link(cfg):
            messagebox.showinfo("提示", "请先配置连接：选串口（COM 口），或切到『无线(WiFi)』并填好 IP。")
            return None
        return cfg

    def _locate_open(self, cfg):
        try:
            link = _open_link_bounded(cfg)
        except Exception as e:
            self.log("打开串口失败：%s" % e)
            messagebox.showwarning("无法连接 Pico", str(e))
            return None
        if not link.ping():
            self.log("Pico 无响应（检查串口、TX/RX 接线或固件）。")
            link.close()
            messagebox.showwarning("无法连接 Pico", "Pico 无响应，请检查串口与接线。")
            return None
        return link

    def _update_pos_label(self, x, y):
        self.ui["loc_pos"].configure(text="当前光标: (%d, %d)" % (x, y))

    def _loc_close_link(self):
        """释放坐标定位期间持有的持久连接。
        在『测试连接 / 自动检测 / 开始运行 / 退出』等需要独占串口前调用，避免端口被占用。"""
        if self._loc_link is not None:
            try:
                self._loc_link.close()
            except Exception:
                pass
            self._loc_link = None

    def _loc_ensure(self, cfg):
        """取一条坐标定位用的持久连接：若已持有且仍在线则复用，否则（重新）打开。
        复用连接可避免每次点击方向键都重开串口、触发板子复位，导致首条指令在启动期丢 ack（表现为『点了没反应』）。"""
        link = self._loc_link
        if link is not None:
            try:
                if link.ping():
                    return link
            except Exception:
                pass
            try:
                link.close()
            except Exception:
                pass
            self._loc_link = None
        link = self._locate_open(cfg)
        self._loc_link = link
        return link

    def _locate_home(self):
        cfg = self._locate_get_cfg()
        if cfg is None:
            return
        if not messagebox.askyesno("归零", "即将把工控机光标甩到屏幕左上角并建立坐标原点。\n（无需手动移动光标，点『是』即执行）"):
            return
        # 注意：归零后不关闭连接，定位期间复用，便于后续方向键/跳转直接生效
        link = self._loc_ensure(cfg)
        if link is None:
            return
        try:
            link.send({"op": "home"})
            self._update_pos_label(0, 0)
            self.log("已归零：光标原点 = 屏幕左上角 (0, 0)。用方向键移动光标，对准后按 F8。")
        except Exception as e:
            self.log("归零失败: %s" % e)

    def _locate_arrow(self, dxn, dyn):
        cfg = self._locate_get_cfg()
        if cfg is None:
            return
        try:
            step = int(self.ui["loc_step"].get() or 10) or 10
        except ValueError:
            self.log("步长不是整数，已使用默认 10。")
            step = 10
        link = self._loc_ensure(cfg)
        if link is None:
            return
        try:
            pos = link.get_pos()
            if pos is None or pos[0] < 0 or pos[1] < 0:
                # 未归零：自动 home 建立原点（等价于点『●归零』），保证操作必有响应
                self.log("尚未建立原点，自动执行归零（home）…")
                link.send({"op": "home"})
                pos = link.get_pos()
            if pos is None or pos[0] < 0 or pos[1] < 0:
                self.log("归零后仍无法获取坐标，请检查设备连接/固件。")
                messagebox.showwarning("坐标定位", "归零后仍未获取到坐标，请检查串口与设备固件。")
                return
            nx, ny = pos[0] + dxn * step, pos[1] + dyn * step
            r = link.send({"op": "move_to", "x": nx, "y": ny})
            if r.get("ack") == "ok":
                self._update_pos_label(nx, ny)
        except Exception as e:
            self.log("移动光标失败: %s" % e)

    def _locate_goto(self):
        """将光标直接移动到输入的绝对坐标，格式为 x,y。"""
        cfg = self._locate_get_cfg()
        if cfg is None:
            return
        raw = self.ui["loc_target"].get().replace("，", ",").replace(" ", "")
        try:
            x, y = [int(v) for v in raw.split(",")]
            if x < 0 or y < 0:
                raise ValueError
        except (ValueError, TypeError):
            messagebox.showwarning("坐标格式", "请输入非负整数坐标，例如：120,96")
            return
        link = self._loc_ensure(cfg)
        if link is None:
            return
        try:
            pos = link.get_pos()
            if pos is None or pos[0] < 0 or pos[1] < 0:
                self.log("尚未建立原点，自动执行归零（home）…")
                link.send({"op": "home"})
                pos = link.get_pos()
            if pos is None or pos[0] < 0 or pos[1] < 0:
                messagebox.showwarning("坐标定位", "归零后仍未获取到坐标，请检查串口与设备固件。")
                return
            resp = link.send({"op": "move_to", "x": x, "y": y})
            if resp.get("ack") == "ok":
                self._update_pos_label(x, y)
                self.log("已跳转到坐标 (%d, %d)" % (x, y))
            else:
                self.log("坐标跳转失败：%s" % resp)
        except Exception as e:
            self.log("坐标跳转失败：%s" % e)

    def _locate_capture(self):
        """F8：把当前光标坐标写入所选步骤（单击/双击/长按/移动）。"""
        cfg = self._locate_get_cfg()
        if cfg is None:
            return
        idx = self._selected_index()
        objlist = self.objects.get(self.current_obj, [])
        s = objlist[idx] if (idx is not None and 0 <= idx < len(objlist)) else None
        if s is None or s.get("type") not in _MOUSE_STEP_SET:
            self.log("F8：请先在『步骤流程』里选中一个鼠标类步骤（单击/双击/长按/移动）。")
            messagebox.showinfo("坐标定位", "请先在「步骤流程」里选中一个鼠标类步骤\n（单击/双击/长按/移动），再按 F8。")
            return
        link = self._loc_ensure(cfg)
        if link is None:
            return
        try:
            pos = link.get_pos()
            if pos is None:
                self.log("尚未建立原点，自动执行归零（home）…")
                link.send({"op": "home"})
                pos = link.get_pos()
            if pos is None:
                self.log("还没建立原点，请先点『●归零』再按 F8。")
                messagebox.showinfo("坐标定位", "请先点『●归零』，把光标挪到目标点后再按 F8。")
                return
            x, y = pos
            s["x"], s["y"] = x, y
            self._refresh_tree()
            self.ui["tree"].selection_set(str(idx))
            self._mark_dirty()
            self.log("已把坐标 (%d, %d) 写入步骤『%s』（记得 Ctrl+S 保存）" % (x, y, s.get("name") or s.get("type")))
        except Exception as e:
            self.log("记录坐标失败: %s" % e)

    # ---------------- worker 线程封装 ----------------
    def _spawn(self, fn, cfg):
        self.stop_event.clear()
        self.confirm_event.set()
        self.current_link = None
        self._loc_close_link()  # 运行前释放定位连接，交给 worker 独占串口
        self.interact = GuiInteract(self.queue, self.stop_event)
        self.worker = threading.Thread(target=self._worker_wrap, args=(fn, cfg), daemon=True)
        self.worker.start()

    def _worker_wrap(self, fn, cfg):
        self.queue.put(("ui", ("running", True)))
        self.log("任务启动")
        try:
            fn(cfg)
        except InterruptedError:
            self.log("任务已停止（用户中断）")
        except Exception as e:
            self.log("任务出错: %s" % e)
            self.log(traceback.format_exc())
        finally:
            self.queue.put(("ui", ("running", False)))
            if not self.stop_event.is_set():
                self.queue.put(("ui", ("beep_done", 1)))
            self.log("任务结束")

    def _confirm_proceed(self):
        self.confirm_event.set()

    def _shortcut_start(self, _ev=None):
        if not self.running:
            self._start_fill()
        return "break"

    def _shortcut_stop(self, _ev=None):
        if self.running:
            self._stop_worker()
        return "break"

    def _shortcut_confirm(self, _ev=None):
        """空格/回车：仅当『已确认』按钮可用且焦点不在输入控件时触发。"""
        try:
            if str(self.ui["btn_confirm"].cget("state")) != "normal":
                return "break"
        except Exception:
            return "break"
        w = self.root.focus_get()
        if w is not None:
            try:
                cls = w.winfo_class()
            except Exception:
                cls = ""
            if cls in ("Entry", "TEntry", "TSpinbox", "TCombobox", "Text",
                       "Listbox", "TListbox", "Treeview"):
                return "break"
        self._confirm_proceed()
        return "break"

    def _key_save(self, _ev=None):
        self._save_config()
        return "break"

    def _stop_worker(self):
        # 只置“停止”信号；不要动 confirm_event，否则进行中的确认会被误判成“已确认”
        self.stop_event.set()
        self.log("正在停止…")
        self._stop_watchdog_rounds = 0
        self._stop_watchdog()

    def _stop_watchdog(self):
        """看门狗：停止请求发出后，若 worker 线程迟迟不结束（如串口卡死），
        12 秒后强制恢复界面，保证『按停止』永远能恢复操作。"""
        self._stop_watchdog_rounds = getattr(self, "_stop_watchdog_rounds", 0) + 1
        if not self.running or not self.stop_event.is_set():
            return
        if self.worker is None or not self.worker.is_alive():
            return  # worker 已结束，就等它的 (running, False) 消息恢复界面
        if self._stop_watchdog_rounds >= 6:  # 6 × 2s = 12s
            self.log("↑ worker 线程卡住（可能串口被占用/设备已拔出），正在强制关闭串口并恢复界面…")
            link = self.current_link
            if link is not None:
                try:
                    link.close()
                    self.log("已强制关闭串口，释放连接。")
                except Exception:
                    pass
                self.current_link = None
            self._apply_ui(("running", False))
            return
        self.root.after(2000, self._stop_watchdog)

    def _on_close(self):
        if self.running and not messagebox.askyesno("退出", "任务正在运行，确定退出吗？"):
            return
        if self.dirty and not messagebox.askyesno("退出", "有未保存的修改（步骤/坐标/按钮），确定退出吗？"):
            return
        self.stop_event.set()
        self.confirm_event.set()
        if self.current_link is not None:
            try:
                self.current_link.close()
            except Exception:
                pass
            self.current_link = None
        self._loc_close_link()
        if self.worker:
            self.worker.join(timeout=3)
        self.root.destroy()

    def _make_confirm(self):
        """返回一个 confirm(label)->bool 的函数：在 GUI 里点亮『已确认』按钮等用户点击。"""
        def confirm(label):
            if self.stop_event.is_set():
                return False
            self.confirm_event.clear()
            self.queue.put(("ui", ("confirm_wait", True)))
            self.queue.put(("ui", ("row_label", str(label))))
            self.queue.put(("ui", ("beep", 1)))
            self.log(">>> %s" % label)
            while not self.confirm_event.wait(0.2):
                if self.stop_event.is_set():
                    self.queue.put(("ui", ("confirm_wait", False)))
                    return False
            self.queue.put(("ui", ("confirm_wait", False)))
            return True
        return confirm

    # ---------------- 校准 ----------------
    def _start_calibrate(self):
        if self.running:
            return
        cfg = self._cfg_or_error()
        if cfg is None:
            return
        self.worker_cfg = copy.deepcopy(cfg)
        self._spawn(self._calibrate_worker, cfg)

    def _calibrate_worker(self, cfg):
        link = None
        try:
            link = _open_link_bounded(cfg)
            self.current_link = link
            if not link.ping():
                self.log("Pico 无响应，校准中止。")
                return
            self.log("Pico 在线 ✅")
            if cfg["start_home"]:
                if not self.interact.ask_yesno("归零校准",
                                               "请先把工控机上的鼠标光标移到『屏幕左上角』，\n然后点『是』继续。"):
                    self.log("用户取消，校准中止。")
                    return
                if self.stop_event.is_set():
                    return
                try:
                    link.send({"op": "home"})
                except TimeoutError:
                    if self.stop_event.is_set():
                        return
                    raise
                self.log("光标已归零到屏幕左上角")
            # 直接遍历主流程 + 各按钮的原始步骤对象（含嵌套按钮展开的覆盖），
            # 校准坐标只改这些原始对象，不会破坏 button 引用结构。
            cal = list(run.iter_target_steps_flat(self.worker_cfg))
            if not cal:
                self.log("当前流程和按钮里没有需要鼠标定位的步骤，无需校准坐标。")
                return
            for idx, (obj_label, t) in enumerate(cal):
                self.log("[%d/%d] 校准(%s): %s  动作 %s  坐标 (%d, %d)"
                         % (idx + 1, len(cal), obj_label,
                            t.get("name") or t.get("type"), t["type"], t["x"], t["y"]))
                while True:
                    link.send({"op": "move_to", "x": t["x"], "y": t["y"]})
                    time.sleep(0.3)
                    ok = self.interact.ask_yesno("校准", "光标已移动到 (%d, %d)。\n是否已对准『%s』？"
                                                 % (t["x"], t["y"], t.get("name") or t.get("type")))
                    if ok:
                        self.log("步骤 %s 坐标确认 (%d, %d)" % (t.get("name") or t["type"], t["x"], t["y"]))
                        break
                    adj = self.interact.ask_string("微调坐标", "输入像素偏移，例如：\n+20,0 表示右移 20\n0,-10 表示上移 10")
                    if adj is None:
                        return
                    try:
                        dx, dy = [int(s) for s in adj.replace(" ", "").split(",")]
                    except ValueError:
                        self.log("格式不对，示例：+20,0 或 0,-10")
                        continue
                    t["x"] += dx
                    t["y"] += dy
                    self.log("新坐标: (%d, %d)" % (t["x"], t["y"]))
            # 校准直接改的是 worker_cfg 里的原始步骤（主流程 + 各按钮），
            # 结构（button 引用等）不变，整体保存即可。
            self._write_config(self.worker_cfg)
            self.queue.put(("ui", ("tree_reload", copy.deepcopy(self.worker_cfg))))
            self.log("校准完成，已写回配置: %s ✅" % self.cfg_path)
        finally:
            if self.current_link is link:
                self.current_link = None
            if link:
                link.close()

    # ---------------- 自动填数 ----------------
    def _validate_start(self, cfg):
        """启动前的统一校验：把问题一次性列出，返回问题列表（空 = 通过）。"""
        problems = []
        if not self._has_link(cfg):
            problems.append("请先配置连接：选串口（COM 口），或切到『无线(WiFi)』并填好 IP。")
        if not cfg["data_file"]:
            problems.append("请先选择数据文件（Excel/CSV）。")
        if not cfg["steps"]:
            problems.append("步骤流程为空，请先添加步骤。")
        for st in cfg["steps"]:
            if st.get("type") == "input" and not st.get("col") and st.get("value") is None and st.get("default") is None:
                problems.append("有一个『输入数值』步骤既没填表格列/固定值/兜底值之一。")
                break
            if st.get("type") == "button" and st.get("name") not in (cfg.get("buttons") or {}):
                problems.append("步骤引用了不存在的自定义按钮：%s" % st.get("name"))
        # 列名校验：字符串列名与表头精确匹配；整数视为 0-based 列序号并检查范围
        if self._col_choices:
            for st in cfg["steps"]:
                if st.get("type") != "input":
                    continue
                c = st.get("col")
                if not c:
                    continue
                if isinstance(c, int):
                    if c < 0 or c >= len(self._col_choices):
                        problems.append("步骤引用的表格列序号 %d 超出表头范围（共 %d 列）。"
                                        % (c, len(self._col_choices)))
                elif c not in self._col_choices:
                    problems.append("步骤引用的表格列『%s』不在表头中。" % c)
        return problems

    def _start_fill(self):
        if self.running:
            return
        cfg = self._cfg_or_error()
        if cfg is None:
            return
        problems = self._validate_start(cfg)
        if problems:
            messagebox.showwarning("请先检查配置", "\n".join("• " + p for p in problems))
            return
        self.worker_cfg = copy.deepcopy(cfg)
        self._spawn(self._fill_worker, cfg)

    def _fill_worker(self, cfg):
        header, records = run.load_records(cfg)
        self.queue.put(("ui", ("row_label", "已读取 %d 行数据" % len(records))))
        self.log("表格：%s  共 %d 行" % (cfg["data_file"], len(records)))
        if not records:
            self.log("表格没有数据行。")
            return

        start_row = 0
        prog = run.load_progress()
        if prog and prog.get("data_file") == cfg["data_file"] and prog.get("row", 0) < len(records):
            ans = self.interact.ask_yesno("断点续传",
                                          "上次已完成 %d/%d 行。\n从断点继续？" % (prog["row"], prog.get("total", len(records))))
            if ans:
                start_row = prog["row"]
                self.log("从断点继续：第 %d 行之后" % (start_row + 1))

        link = _open_link_bounded(cfg)
        try:
            self.current_link = link
            if not link.ping():
                self.log("Pico 无响应，任务中止。")
                return
            self.log("Pico 在线 ✅")
            self._run_rows(link, cfg, records, start_row, header)
        finally:
            if self.current_link is link:
                self.current_link = None
            link.close()

    def _auto_home(self, link, cfg, tag="", force=False):
        """自动归零到屏幕左上角原点。

        - 流程每一步执行前归零：跟随 start_home 开关（force=False），保证
          每一步坐标都从 (0,0) 精确出发，不受上一步真实鼠标移动或人工操作
          造成的基准偏移影响。
        - 人工确认后归零：force=True，无条件执行。因为确认期间用户可能移动
          工控机真实鼠标，相对鼠标会在系统光标上叠加位移，使内部坐标基准偏移，
          必须重新归零才能保证后续坐标动作从 (0,0) 精确出发。
        """
        if cfg.get("start_home") or force:
            if self.stop_event.is_set():
                return False
            try:
                link.send({"op": "home"})
            except TimeoutError:
                if self.stop_event.is_set():
                    return False
                raise
            self.log("光标已归零到屏幕左上角%s" % tag)
        return True

    def _run_rows(self, link, cfg, records, start_row, header=None):
        steps = run.expand_steps(cfg)
        delay = int(cfg.get("step_delay_ms", 300))
        total = len(records)
        # 填充范围：row_start/row_end（1-based 数据行），并与断点 start_row 取交集
        sel_start = int(cfg.get("row_start", 1) or 1)
        sel_end = int(cfg.get("row_end", 0) or 0)
        begin = max(start_row, sel_start - 1)
        end = total if sel_end <= 0 else min(total, sel_end)
        if begin >= end:
            self.log("没有待处理的行（已超出所选范围或已完成）。")
            return
        confirm_fn = self._make_confirm()
        self.queue.put(("ui", ("bar_max", max(1, end - begin))))
        self.log("本次将填充数据行：第 %d–%d 行（共 %d 行）。" % (begin + 1, end, total))
        for i in range(begin, end):
            if self.stop_event.is_set():
                return
            rec = records[i]
            cmds, summary = run.build_record_commands(steps, rec, delay, header)
            self.queue.put(("ui", ("row_label", "第 %d/%d 行" % (i + 1, end))))
            self.queue.put(("ui", ("summary", "   ".join("[%s] %s" % (k, v) for k, v in summary.items()))))
            self.log("── 第 %d/%d 行 ──" % (i + 1, end))
            for k, v in summary.items():
                self.log("  %-12s %s" % (k, v))
            for c in cmds:
                if self.stop_event.is_set():
                    return
                if c["op"] == "__confirm__":
                    if not confirm_fn(c.get("label", "确认")):
                        return
                    # 人工确认期间用户可能移动工控机真实鼠标，使系统光标基准偏移；
                    # 确认后强制归零，保证后续坐标动作从 (0,0) 精确出发。
                    if not self._auto_home(link, cfg, "（人工确认后）", force=True):
                        return
                    continue
                if c["op"] == "sleep":
                    if not _interruptible_sleep(c["ms"], self.stop_event):
                        return
                    continue
                # 每一步流程执行前自动归零：保证该步坐标从屏幕左上角原点 (0,0)
                # 精确出发，不受上一步真实鼠标移动或人工操作造成的基准偏移影响。
                if not self._auto_home(link, cfg, "（步骤前）"):
                    return
                try:
                    resp = link.send(c)
                except TimeoutError:
                    if self.stop_event.is_set():
                        return
                    raise
                if resp.get("ack") == "error":
                    raise RuntimeError("Pico 执行失败: %s （指令: %s）" % (resp.get("msg"), c))
            run.save_progress(i + 1, total, cfg["data_file"])
            self.queue.put(("ui", ("bar", i + 1 - begin)))
            if cfg["confirm"] == "enter":
                if not confirm_fn(">>> 请到工控机屏幕上确认本行数值无误，然后点『已确认』继续"):
                    return
        run.clear_progress()
        self.log("全部完成 ✅  本次共 %d 行。" % (end - begin))

    # ---------------- 单步测试 ----------------
    def _test_selected_step(self):
        """立即执行步骤列表中当前选中的单个步骤，不读取 Excel。"""
        if self.running:
            messagebox.showinfo("提示", "正在执行其他任务，请稍候。")
            return
        idx = self._selected_index()
        steps = self.objects.get(self.current_obj) or []
        if idx is None or not (0 <= idx < len(steps)):
            messagebox.showinfo("单步测试", "请先选中步骤流程中的一行。")
            return
        step = copy.deepcopy(steps[idx])
        typ = step.get("type", "click")
        if typ == "button":
            messagebox.showinfo("单步测试", "按钮步骤不能单独测试，请测试按钮中的具体步骤。")
            return
        cfg = self._cfg_or_error()
        if cfg is None or not self._has_link(cfg):
            return
        if not messagebox.askyesno("单步测试", "立即执行步骤『%s』？\n光标可能会移动或产生按键输入。"
                                   % (step.get("name") or STEP_TYPE_LABELS.get(typ, typ))):
            return
        # 单步测试不叠加默认延时，保证“立即测试”。
        step.pop("delay_ms", None)
        cfg2 = dict(self._link_cfg(cfg),
                    step_delay_ms=0, confirm="none", steps=[step],
                    buttons={})
        self.worker_cfg = copy.deepcopy(cfg2)
        self._spawn(self._test_step_worker, cfg2)

    def _test_step_worker(self, cfg):
        link = None
        try:
            link = _open_link_bounded(cfg)
            self.current_link = link
            if not link.ping():
                self.log("单步测试失败：Pico 无响应。")
                return
            step = cfg["steps"][0]
            typ = step.get("type", "click")
            # 坐标型动作：仅当尚未建立原点时才补一次 home（不每次重置），保持坐标连续。
            if typ in _MOUSE_STEP_SET:
                pos = link.send({"op": "get_pos"})
                if pos.get("ack") != "ok":
                    resp = link.send({"op": "home"})
                    if resp.get("ack") != "ok":
                        raise RuntimeError("归零失败：%s" % resp)
            cmds, _ = run.build_record_commands(cfg["steps"], {}, 0)
            for cmd in cmds:
                if self.stop_event.is_set():
                    return
                if cmd["op"] == "sleep":
                    continue
                resp = link.send(cmd)
                if resp.get("ack") == "error":
                    raise RuntimeError("Pico 执行失败：%s" % resp.get("msg"))
            self.log("单步测试完成：%s ✅" % (step.get("name") or STEP_TYPE_LABELS.get(typ, typ)))
        except Exception as e:
            self.log("单步测试失败：%s" % e)
        finally:
            if self.current_link is link:
                self.current_link = None
            if link:
                link.close()

    # ---------------- 按钮快捷执行 ----------------
    def _run_button(self, name):
        if self.running:
            messagebox.showinfo("提示", "正在执行其他任务，请稍候。")
            return
        cfg = self._cfg_or_error()
        if cfg is None:
            return
        btn_steps = self.objects.get("btn:" + name)
        if not btn_steps:
            messagebox.showwarning("提示", "按钮不存在：%s" % name)
            return
        cfg2 = dict(self._link_cfg(cfg),
                    start_home=cfg["start_home"], step_delay_ms=cfg["step_delay_ms"],
                    confirm="none", steps=copy.deepcopy(btn_steps),
                    buttons=copy.deepcopy(cfg.get("buttons") or {}))
        self.worker_cfg = copy.deepcopy(cfg2)
        self._spawn(self._button_worker, cfg2)

    def _button_worker(self, cfg):
        link = None
        try:
            link = _open_link_bounded(cfg)
            self.current_link = link
            if not link.ping():
                self.log("Pico 无响应，按钮未执行。")
                return
            self.log("Pico 在线 ✅")
            steps = run.expand_steps(cfg)
            cmds, summary = run.build_record_commands(steps, {}, cfg["step_delay_ms"])
            self.log("执行按钮，共 %d 条指令" % len(cmds))
            for k, v in summary.items():
                self.log("  %-12s %s" % (k, v))
            confirm_fn = self._make_confirm()
            for c in cmds:
                if self.stop_event.is_set():
                    return
                if c["op"] == "__confirm__":
                    if not confirm_fn(c.get("label", "确认")):
                        return
                    # 人工确认期间用户可能移动工控机真实鼠标，使系统光标基准偏移；
                    # 确认后强制归零，保证后续坐标动作从 (0,0) 精确出发。
                    if not self._auto_home(link, cfg, "（人工确认后）", force=True):
                        return
                    continue
                if c["op"] == "sleep":
                    if not _interruptible_sleep(c["ms"], self.stop_event):
                        return
                    continue
                # 每一步流程执行前自动归零：保证该步坐标从屏幕左上角原点 (0,0)
                # 精确出发，不受上一步真实鼠标移动或人工操作造成的基准偏移影响。
                if not self._auto_home(link, cfg, "（步骤前）"):
                    return
                try:
                    resp = link.send(c)
                except TimeoutError:
                    if self.stop_event.is_set():
                        return
                    raise
                if resp.get("ack") == "error":
                    raise RuntimeError("Pico 执行失败: %s （指令: %s）" % (resp.get("msg"), c))
            self.log("按钮执行完成 ✅")
        except Exception as e:
            self.log("按钮执行失败: %s" % e)
        finally:
            if self.current_link is link:
                self.current_link = None
            if link:
                link.close()

    # ---------------- Dry Run ----------------
    def _dry_run(self):
        cfg = self._cfg_or_error()
        if cfg is None:
            return
        try:
            header, records = run.load_records(cfg)
        except Exception as e:
            messagebox.showerror("读取表格失败", str(e))
            return
        self.log("== DRY RUN：仅预览指令，不发送 ==")
        delay = int(cfg.get("step_delay_ms", 300))
        steps = run.expand_steps(cfg)
        sel_start = int(cfg.get("row_start", 1) or 1)
        sel_end = int(cfg.get("row_end", 0) or 0)
        begin = max(0, sel_start - 1)
        end = len(records) if sel_end <= 0 else min(len(records), sel_end)
        self.log("范围：数据行第 %d–%d 行，共 %d 行。"
                 % (begin + 1, end, max(0, end - begin)))
        for i, rec in enumerate(records[begin:end]):
            cmds, summary = run.build_record_commands(steps, rec, delay, header)
            if i >= 2:
                self.log("… 其余 %d 行类似" % (max(0, end - begin) - 2))
                break
            self.log("-- 第 %d 行 --" % (begin + i + 1))
            for c in cmds:
                self.log("    %s" % json.dumps(c, ensure_ascii=False))


def main():
    if sys.platform == "win32":
        # 让 Tk 感知系统 DPI，避免高分屏下文字发虚
        try:
            import ctypes
            try:
                ctypes.windll.shcore.SetProcessDpiAwareness(1)
            except Exception:
                ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass
    root = tk.Tk()
    if not _ensure_run():
        root.withdraw()
        messagebox.showerror("缺少依赖",
                             "需要 pyserial（串口）和 openpyxl（Excel）。\n\n"
                             "请先执行：\n    pip install pyserial openpyxl")
        root.destroy()
        return
    root.title("工控机自动填数助手")
    # 高度放大到 1000：连接面板新增「无线控制」3 行（约 100px），
    # 而「按钮快捷执行/进度/日志」三个区块（合计约 520px）是运行时才映射的，
    # 沿用 860 会让日志区被挤出可视范围。屏幕为 2560x1440，余量充足。
    root.geometry("1400x1000")
    root.minsize(1100, 760)
    root.report_callback_exception = lambda exc, val, tb: _dump_crash("callback", exc, val, tb, root)
    try:
        App(root)
        # 默认最大化：整个界面按 pack 顺序纵向排列，内容总高约 1700px，
        # 高度不足时 Tk 会直接不映射（丢弃）放不下的底部区块（按钮栏/操作栏/进度/日志）。
        # 最大化后可用高度约 1392px（2560x1440 屏），可保证底部四块都在。
        try:
            root.state("zoomed")
        except Exception:
            try:
                root.attributes("-zoomed", True)
            except Exception:
                pass
        root.mainloop()
    except Exception as e:
        _dump_crash("mainloop", type(e), e, e.__traceback__, root)
        raise


if __name__ == "__main__":
    main()
