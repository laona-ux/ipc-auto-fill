#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
笔记本端主程序：读 Excel/CSV 表格，把数值一条条自动填入工控机（经 Pico HID 注入）。

用法：
    python run.py                          # 正常跑（config.json）
    python run.py --calibrate              # 坐标校准模式：逐个移动光标让用户确认
    python run.py --dry-run                # 只打印将要发送的指令序列，不连串口
    python run.py --resume                 # 直接从断点继续（不询问）
    python run.py --config my.json         # 指定配置文件

依赖：pip install pyserial openpyxl
"""
import argparse
import csv
import hashlib
import json
import os
import sys
import time

try:
    import serial
except ImportError:
    sys.exit("缺少 pyserial，请先执行: pip install pyserial\n"
             "（配置文件读取还需要 openpyxl，一起装: pip install pyserial openpyxl）")

import socket

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None  # 只在用到 .xlsx 时才报错

HERE = os.path.dirname(os.path.abspath(__file__))
PROGRESS_FILE = os.path.join(HERE, ".progress")
AUDIT_DIR = os.path.join(HERE, "audit")
TIMEOUT = 5.0          # 单条指令等待应答的超时（秒）
PING_TIMEOUT = 8.0     # 开机自检的宽松超时

# type 文本可安全注入的最大长度（超出截断）
MAX_TYPE_LEN = 64
# key 步骤允许的按键白名单（含修饰键组合写法）；刻意不放 win / alt，防误注入危险组合键。
ALLOWED_KEYS = {
    "enter", "tab", "esc", "escape", "f1", "f2", "f3", "f4", "f5", "f6",
    "f7", "f8", "f9", "f10", "f11", "f12", "backspace", "delete", "home",
    "end", "pageup", "pagedown", "up", "down", "left", "right", "insert",
    "space", "ctrl+a", "ctrl+c", "ctrl+v", "ctrl+x", "ctrl+z", "ctrl+y",
    "ctrl+s", "shift+tab", "shift+enter", "ctrl+tab",
}
# 加白允许的单字符（字母/数字/常用符号，Pico 可打出的 ASCII 导航/字符键）
ALLOWED_SINGLE_CHARS = set("abcdefghijklmnopqrstuvwxyz0123456789")

# 审计日志字段清单（新字段需补充 AUditSchema 校验）
AUDIT_KEYS = ("time", "run", "row", "total", "data_file", "summary",
              "confirm_required", "confirmed", "status")


# ---------------- 工具函数 ----------------

def read_config(path):
    with open(path, "r", encoding="utf-8-sig") as f:
        cfg = json.load(f)
    if not cfg.get("targets") and not cfg.get("steps"):
        raise ValueError("config 里 targets 和 steps 都为空，请至少配置一个输入框目标或步骤流程")
    return cfg


def _rows_to_records(rows, header_row, has_header):
    """把原始行数据 (list[tuple]) 转成 (表头列表, 行列表[dict])。
    header_row: 表头所在行号(0-based)；has_header=false 时不读表头，自动生成 col1..colN。
    无表头时 header_row 表示第一条数据所在行号，可用于跳过前面的标题行。"""
    if header_row < 0 or header_row >= len(rows):
        raise ValueError("header_row=%d 越界（表格共 %d 行）" % (header_row, len(rows)))
    if has_header:
        header = [str(c).strip() if c is not None else "" for c in rows[header_row]]
        header = [h if h else "col%d" % (i + 1) for i, h in enumerate(header)]
        data_start = header_row + 1
    else:
        ncol = max((len(r) for r in rows[header_row:]), default=0)
        header = ["col%d" % (i + 1) for i in range(ncol)]
        data_start = header_row
    records = []
    for r in rows[data_start:]:
        rec = {}
        for i, h in enumerate(header):
            rec[h] = r[i] if i < len(r) else None
        if any(v is not None and str(v).strip() != "" for v in rec.values()):
            records.append(rec)
    return header, records


def load_records(cfg):
    """读表格，返回 (表头列表, 行列表[dict])。支持 .xlsx/.xlsm/.xls 和 .csv。
    可用配置项：header_row（表头行，默认0）、has_header（默认 true）。"""
    path = cfg["data_file"]
    if not os.path.isabs(path):
        path = os.path.join(HERE, path)
    if not os.path.exists(path):
        raise FileNotFoundError("找不到表格文件: %s" % path)
    header_row = int(cfg.get("header_row", 0) or 0)
    has_header = bool(cfg.get("has_header", True))
    ext = os.path.splitext(path)[1].lower()

    if ext in (".xlsx", ".xlsm"):
        if load_workbook is None:
            raise ImportError("读取 xlsx 需要 openpyxl，请执行: pip install openpyxl")
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb[cfg.get("sheet")] if cfg.get("sheet") else wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            raise ValueError("表格是空白的")
        return _rows_to_records(rows, header_row, has_header)
    elif ext == ".xls":
        try:
            import xlrd
        except ImportError:
            raise ImportError("读取 .xls 需要 xlrd，请执行: pip install xlrd")
        book = xlrd.open_workbook(path)
        ws = book.sheet_by_name(cfg.get("sheet")) if cfg.get("sheet") else book.sheet_by_index(0)
        rows = [tuple(r) for r in ws.get_rows()]
        if not rows:
            raise ValueError("表格是空白的")
        return _rows_to_records(rows, header_row, has_header)
    elif ext == ".csv":
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            rows = [r for r in csv.reader(f)]
        if not rows:
            raise ValueError("表格是空白的")
        return _rows_to_records(rows, header_row, has_header)
    else:
        raise ValueError("不支持的表格类型: %s（支持 .xlsx/.xlsm/.xls/.csv）" % ext)


def fmt_value(v, round_dp=None):
    """把单元格值转成输入文本：1500.0 -> '1500'，25.5 -> '25.5'。
    若 round_dp 指定（保留 N 位小数），按固定小数位输出（如 6.14054757143477 -> '6.14'）。"""
    if v is None:
        return None
    if isinstance(v, bool):
        v = int(v)
    if isinstance(v, float) and round_dp is not None:
        return f"{v:.{int(round_dp)}f}"
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


# ---------------- 文本净化 / 按键白名单（安全加固） ----------------

def _sanitize_text(txt, max_len=MAX_TYPE_LEN):
    """把将要 type 的文本净化：
    - 去掉/替换控制字符（尤其 \\n、\\t 会被 Pico 固件转成真实回车/Tab 触发计划外按键）；
    - 截断到 max_len。
    返回净化后的字符串；若净化为空则原样返回（由上层决定是否跳过）。"""
    if txt is None:
        return None
    s = str(txt)
    # 控制字符（含 \\r \\n \\t \\x00-\\x1f、\\x7f）一律替换为空格，避免注入真实按键
    s = "".join(ch if ch >= " " and ch != "\x7f" else " " for ch in s)
    if max_len and len(s) > max_len:
        s = s[:max_len]
    return s


def validate_key(key):
    """校验一个 key 步骤的按键名是否在白名单内。
    合法返回 (True, None)；非法返回 (False, 理由)。用于防危险组合键（win/alt）误注入。"""
    k = str(key or "").strip().lower()
    if k in ALLOWED_KEYS:
        return True, None
    if len(k) == 1 and k in ALLOWED_SINGLE_CHARS:
        return True, None
    # 允许 "win+..." 之外的合法修饰组合（默认只放行 ctrl/shift 组合，且不含 win/alt）
    if "+" in k:
        parts = k.split("+")
        if any(p in ("win", "alt", "cmd", "option") for p in parts):
            return False, "按键含 win/alt 修饰键，已被安全白名单拦截: %s" % k
        if all(p in ALLOWED_KEYS or (len(p) == 1 and p in ALLOWED_SINGLE_CHARS) or p in ("ctrl", "shift")
               for p in parts):
            return True, None
    return False, "按键不在允许范围内: %s" % k


# ---------------- 串口通信 ----------------

class Link:
    def __init__(self, port, baud):
        # 短超时配合按行读取，避免每条短 JSON 指令额外等待 1 秒。
        self.ser = serial.Serial(port, baud, timeout=0.05, write_timeout=1)

    def close(self):
        try:
            self.ser.close()
        except Exception:
            pass

    def send(self, cmd, timeout=TIMEOUT):
        """发一条指令，等 Pico 回 ack（ping 返回 pong）。返回应答 dict。"""
        self.ser.reset_input_buffer()
        line = json.dumps(cmd, ensure_ascii=False) + "\n"
        self.ser.write(line.encode("utf-8"))
        self.ser.flush()
        buf = b""
        deadline = time.time() + timeout
        while time.time() < deadline:
            # Pico 每条应答以换行结束；按行读取可在收到完整应答后立即返回。
            data = self.ser.read_until(b"\n")
            if data:
                buf += data
                while b"\n" in buf:
                    raw, buf = buf.split(b"\n", 1)
                    raw = raw.strip()
                    if not raw:
                        continue
                    try:
                        obj = json.loads(raw.decode("utf-8", errors="replace"))
                    except ValueError:
                        continue
                    if "pong" in obj or "ack" in obj:
                        return obj
        raise TimeoutError("指令无应答: %s" % cmd)

    def ping(self):
        for _ in range(3):
            try:
                return self.send({"op": "ping"}, timeout=PING_TIMEOUT).get("pong") is True
            except TimeoutError:
                time.sleep(0.5)
        return False

    def get_pos(self, timeout=TIMEOUT):
        """查询 Pico 当前跟踪的光标坐标 (x, y)；未归零或失败返回 None。"""
        resp = self.send({"op": "get_pos"}, timeout=timeout)
        if resp.get("ack") == "ok":
            return int(resp["x"]), int(resp["y"])
        return None


class TcpLink:
    """与 Link 接口完全一致，但通过 TCP 连接 ESP32 的 WiFi TCP Server。

    协议不变：每行一条 JSON（\\n 结尾），板子回一行 {"ack":...} / {"pong":true}。
    串口是「字节流 + reset_input_buffer」；TCP 是「字节流 + 读前丢弃残留」，二者语义对齐。
    """

    def __init__(self, host, port, connect_timeout=5.0, timeout=0.05):
        self.sock = socket.create_connection((host, int(port)), timeout=connect_timeout)
        self.sock.settimeout(timeout)
        self._buf = b""

    def close(self):
        try:
            self.sock.close()
        except Exception:
            pass

    def _drain(self):
        """读前丢弃缓冲区里残留的、上一次未读完的响应（对应串口 reset_input_buffer）。"""
        self.sock.settimeout(0)
        try:
            while True:
                try:
                    d = self.sock.recv(4096)
                except (socket.timeout, BlockingIOError, OSError):
                    break
                if not d:
                    break
        except Exception:
            pass

    def send(self, cmd, timeout=TIMEOUT):
        self._drain()
        self.sock.settimeout(timeout)
        line = json.dumps(cmd, ensure_ascii=False) + "\n"
        self.sock.sendall(line.encode("utf-8"))
        buf = b""
        deadline = time.time() + timeout
        while time.time() < deadline:
            try:
                data = self.sock.recv(4096)
            except socket.timeout:
                data = b""
            if not data:
                break
            buf += data
            while b"\n" in buf:
                raw, buf = buf.split(b"\n", 1)
                raw = raw.strip()
                if not raw:
                    continue
                try:
                    obj = json.loads(raw.decode("utf-8", errors="replace"))
                except ValueError:
                    continue
                if "pong" in obj or "ack" in obj:
                    return obj
        raise TimeoutError("指令无应答: %s" % cmd)

    def ping(self):
        for _ in range(3):
            try:
                return self.send({"op": "ping"}, timeout=PING_TIMEOUT).get("pong") is True
            except TimeoutError:
                time.sleep(0.5)
        return False

    def get_pos(self, timeout=TIMEOUT):
        """查询板子当前跟踪的光标坐标 (x, y)；未归零或失败返回 None。"""
        resp = self.send({"op": "get_pos"}, timeout=timeout)
        if resp.get("ack") == "ok":
            return int(resp["x"]), int(resp["y"])
        return None


def open_link(cfg):
    """按 transport 选择串口或 WiFi(TCP)。cfg 需含 transport / com_port / baudrate / wifi_ip / wifi_port。"""
    transport = (cfg.get("transport") or "serial").lower()
    if transport == "wifi":
        return TcpLink(cfg["wifi_ip"], int(cfg.get("wifi_port", 8080)))
    return Link(cfg["com_port"], cfg.get("baudrate", 115200))


# ---------------- 指令序列 / 步骤宏 ----------------
# 一套「步骤流程」（steps/宏）：每一行表格记录按步骤顺序执行一遍。
# 支持步骤类型：
#   click   移动到 (x,y) 并单击（进入界面、点击按钮）
#   dblclick 移动到 (x,y) 并双击
#   press_at 移动到 (x,y) 并按住 ms 毫秒（长按）
#   move    只移动到 (x,y)，不点击
#   scroll  滚动滚轮，delta>0 上滚 / <0 下滚
#   key     按一个键，可加 times 重复（enter/tab/esc/f5/ctrl+a/shift+tab …）
#   input   输入一个数值：来源 = 表格列 col 或固定值 value；可先按 pre_keys、回车
#   sleep   等待 ms 毫秒
#   home    让 Pico 把光标甩回屏幕左上角并重设原点
#   confirm 中途等待人工确认（GUI 弹按钮 / CLI 回车）
#   button  引用自定义按钮（config 里 buttons 下的某段步骤，执行时展开）
# 每个步骤可选 name 作为它在日志/汇总里的说明。

class UserQuit(Exception):
    """用户主动退出，保留断点。"""


def _row_value(record, header, col):
    """按 col 取该行的一个值。col 可以是列名字符串，或 0-based 列序号。"""
    if isinstance(col, int):
        if header and 0 <= col < len(header):
            return record.get(header[col])
        vals = list(record.values())
        return vals[col] if 0 <= col < len(vals) else None
    return record.get(col)


def step_value(step, record, header=None):
    """解析一个 input 步骤要填入的文本。
    优先级：固定值 value > 表格列 col；可再加 default(单元格为空时兜底)、prefix/suffix(前后缀)。"""
    if step.get("value") is not None:
        base = fmt_value(step["value"], step.get("round"))
    else:
        col = step.get("col")
        if col is None or col == "":
            base = None
        else:
            base = fmt_value(_row_value(record, header, col), step.get("round"))
    if base is None or base == "":
        if step.get("default") is not None:
            base = fmt_value(step["default"], step.get("round"))
        else:
            return None
    return "%s%s%s" % (step.get("prefix") or "", base, step.get("suffix") or "")


def _step_cmds(step, record, step_delay):
    """把一个步骤展开成 Pico 指令列表。返回 list[dict(op=...)]。"""
    typ = step.get("type", "click")
    cmds = []
    post = int(step.get("delay_ms", step_delay))   # 动作后的停顿
    if typ == "click":
        cmds.append({"op": "click_at", "x": int(step["x"]), "y": int(step["y"])})
    elif typ == "dblclick":
        cmds.append({"op": "dblclick_at", "x": int(step["x"]), "y": int(step["y"])})
    elif typ == "move":
        cmds.append({"op": "move_to", "x": int(step["x"]), "y": int(step["y"])})
    elif typ == "press_at":
        cmds.append({"op": "press_at", "x": int(step["x"]), "y": int(step["y"]),
                     "ms": int(step.get("ms", 500))})
    elif typ in ("scroll", "wheel"):
        cmds.append({"op": "wheel", "delta": int(step.get("delta", -120))})
    elif typ == "key":
        ok, reason = validate_key(step["key"])
        if not ok:
            raise ValueError(reason)   # 非法按键在生成阶段就拒绝，避免运行时注入
        times = max(1, int(step.get("times", 1)))
        cmds += [{"op": "key", "key": step["key"]}] * times
    elif typ == "input":
        # 值为空且无兜底时整步跳过（连 pre_keys/回车/后延时都不发），
        # 避免把工位上已有的数值清掉（例如空单元格仍发送 Ctrl+A / 回车）。
        val = step_value(step, record)
        if val is None:
            return []
        for k in step.get("pre_keys", []):
            ok, reason = validate_key(k)
            if not ok:
                raise ValueError(reason)
            cmds.append({"op": "key", "key": k})
        cmds.append({"op": "type", "text": _sanitize_text(val)})
        if step.get("enter"):
            cmds.append({"op": "key", "key": "enter"})
    elif typ == "home":
        cmds.append({"op": "home"})
    elif typ == "sleep":
        return [{"op": "sleep", "ms": int(step.get("ms", 0))}]
    # 步骤结束后补一个停顿（让工控机软件有消化时间）；sleep 步骤不再叠加
    if post > 0:
        cmds.append({"op": "sleep", "ms": post})
    return cmds


def _step_desc(step):
    """生成导航类步骤的简短说明（用于日志/每行汇总）。"""
    typ = step.get("type")
    if typ in ("click", "dblclick", "move"):
        act = {"click": "点击", "dblclick": "双击", "move": "移动"}[typ]
        return "(%s %s,%s)" % (act, step.get("x"), step.get("y"))
    if typ == "press_at":
        return "(长按 %s,%s)" % (step.get("x"), step.get("y"))
    if typ in ("scroll", "wheel"):
        d = int(step.get("delta", -120))
        return "(滚轮%s)" % ("上" if d > 0 else "下")
    if typ == "key":
        t = max(1, int(step.get("times", 1)))
        return "(%s%s)" % (step.get("key", "").upper(), ("×%d" % t) if t > 1 else "")
    if typ == "home":
        return "(归零)"
    return ""


def build_record_commands(steps, record, step_delay, header=None):
    """把一行记录按步骤流程展开成指令序列；返回 (命令列表, 汇总 dict)。
    其中 confirm 步骤会用伪指令 {"op":"__confirm__"} 表示，由发送方等待人工确认。
    header：表头列表，供 input 步骤按列序号(col 为整数)取值。"""
    cmds = []
    summary = {}
    for step in steps:
        typ = step.get("type", "click")
        name = step.get("name", "")
        if typ == "confirm":
            cmds.append({"op": "__confirm__", "label": name or "确认"})
            continue
        if typ == "button":
            continue  # 已在 expand_steps 里展开，若未展开则忽略该引用
        cmds.extend(_step_cmds(step, record, step_delay))
        if typ == "input":
            if name:
                val = step_value(step, record, header)
                summary[name] = val if val is not None else "(空，跳过)"
        elif typ in ("click", "dblclick", "move", "press_at", "scroll", "wheel", "key", "home"):
            if name:
                summary[name] = _step_desc(step)
    return cmds, summary


def _steps_from_targets(targets):
    """把旧的 targets 列表转换成等价的步骤流程（兼容旧配置）。"""
    steps = []
    for t in targets:
        name = t.get("name", "")
        if t.get("mode") == "keyboard":
            steps.append({"type": "key", "key": "tab", "times": max(1, int(t.get("tab_to", 1))),
                          "name": name})
        else:
            steps.append({"type": "click", "x": int(t["x"]), "y": int(t["y"]), "name": name})
        for k in t.get("pre_keys", []):
            steps.append({"type": "key", "key": k})
        steps.append({"type": "input", "col": t["col"], "round": t.get("round"), "name": name})
    return steps


def _expand_buttons(steps, buttons, seen=None, depth=0, max_depth=4):
    """把流程里的 button 引用展开成按钮定义的步骤序列（防循环）。"""
    if seen is None:
        seen = set()
    out = []
    for step in steps:
        if step.get("type") != "button":
            out.append(step)
            continue
        bname = step.get("name")
        blist = (buttons or {}).get(bname)
        if not blist or depth >= max_depth or bname in seen:
            out.append(step)   # 无法展开时保留原引用
            continue
        seen.add(bname)
        out.extend(_expand_buttons(blist, buttons, seen, depth + 1, max_depth))
        seen.discard(bname)
    return out


def expand_steps(cfg):
    """返回最终生效的步骤列表：优先 steps，否则由 targets 转换；并展开自定义按钮。"""
    raw = cfg.get("steps")
    if not raw and cfg.get("targets"):
        raw = _steps_from_targets(cfg["targets"])
    if not raw:
        raw = []
    return _expand_buttons(raw, cfg.get("buttons") or {})


# 兼容接口：旧的按表格列逐项填入逻辑（等价于 click/keyboard + input 的步骤序列）。
def build_row_commands(targets, record, step_delay_ms):
    return build_record_commands(_steps_from_targets(targets), record, step_delay_ms)


# ---------------- 断点 / 进度 ----------------

def _cfg_digest(cfg):
    """计算“断点归属”的稳定指纹：步骤流程 + 表格解析参数 + 数据文件信息。
    用于检测两次运行之间配置是否漂移——漂移后禁止（或提示）沿用旧断点。"""
    steps_core = []
    for s in expand_steps(cfg):
        # 只保留影响“填到哪/填什么”的字段，去掉 name/delay 等展示性字段
        sc = {k: s[k] for k in ("type", "col", "round", "key", "x", "y", "ms", "delta")
              if k in s and s[k] is not None}
        steps_core.append(sc)
    parts = [json.dumps(steps_core, sort_keys=True, ensure_ascii=False),
             str(cfg.get("sheet", "")),
             str(int(cfg.get("header_row", 0) or 0)),
             str(bool(cfg.get("has_header", True))),
             str(int(cfg.get("row_start", 1) or 1)),
             str(int(cfg.get("row_end", 0) or 0))]
    data_path = cfg.get("data_file", "")
    if not os.path.isabs(data_path):
        data_path = os.path.join(HERE, data_path)
    try:
        st = os.stat(data_path)
        parts.append("%d_%d" % (int(st.st_mtime), int(st.st_size)))
    except OSError:
        parts.append("")
    return hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()[:16]


def load_progress():
    if os.path.exists(PROGRESS_FILE):
        try:
            with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except (ValueError, OSError):
            pass
    return None


def save_progress(row_index, total, data_file, digest=None):
    data = {"row": row_index, "total": total, "data_file": data_file,
            "time": time.strftime("%Y-%m-%d %H:%M:%S")}
    if digest:
        data["steps_digest"] = digest
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f)


def progress_is_drifted(prog, cfg):
    """断点与当前配置是否“漂移”（步骤/范围/表格解析参数或数据文件已变）。
    True 表示沿用该断点有填错位置的风险，应提示用户。"""
    if not prog:
        return False
    got = prog.get("steps_digest")
    if not got:
        return False  # 旧格式无指纹，保守起见不拦截，仅提示由调用方决定
    return got != _cfg_digest(cfg)


def clear_progress():
    if os.path.exists(PROGRESS_FILE):
        os.remove(PROGRESS_FILE)


# ---------------- 审计日志 ----------------

def _audit_path():
    os.makedirs(AUDIT_DIR, exist_ok=True)
    return os.path.join(AUDIT_DIR, "fill_%s.jsonl" % time.strftime("%Y%m%d"))


def append_audit(rec):
    """把一次“行级填入”的审计记录追加写入 host/audit/fill_YYYYMMDD.jsonl。
    记录：时间、行号、汇总、是否要求确认、确认结果、状态。"""
    try:
        os.makedirs(AUDIT_DIR, exist_ok=True)
        with open(_audit_path(), "a", encoding="utf-8") as f:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")
    except OSError:
        pass  # 审计写失败不阻断主流程


# ---------------- 配置校验 ----------------

def validate_config(cfg, header=None, records=None):
    """返回配置问题列表（空列表 = 合法）。
    覆盖：步骤类型枚举、key 白名单、button 引用存在、x/y 为整数、
    input 列有效（列名在表头 / 整数列号在范围）、范围合法、confirm 枚举。"""
    problems = []
    steps = cfg.get("steps")
    if not steps and cfg.get("targets"):
        steps = _steps_from_targets(cfg["targets"])
    steps = steps or []

    step_types = {"click", "dblclick", "move", "press_at", "scroll", "wheel",
                  "key", "input", "sleep", "home", "confirm", "button"}
    buttons = cfg.get("buttons") or {}
    for idx, s in enumerate(steps):
        tag = "steps[%d]" % idx
        typ = s.get("type", "click")
        if typ not in step_types:
            problems.append("%s: 未知步骤类型 %r" % (tag, typ))
            continue
        if typ in ("click", "dblclick", "move", "press_at"):
            for kk in ("x", "y"):
                v = s.get(kk)
                if not isinstance(v, int) or isinstance(v, bool):
                    try:
                        int(v)
                    except (TypeError, ValueError):
                        problems.append("%s: %s 必须是整数，当前 %r" % (tag, kk, v))
            if typ == "press_at" and "ms" in s:
                try:
                    int(s["ms"])
                except (TypeError, ValueError):
                    problems.append("%s: ms 必须是整数" % tag)
        elif typ == "key":
            ok, why = validate_key(s.get("key"))
            if not ok:
                problems.append("%s: %s" % (tag, why))
        elif typ == "button":
            if s.get("name") not in buttons:
                problems.append("%s: 引用不存在的按钮 %r" % (tag, s.get("name")))
        elif typ == "input" and s.get("col") is not None and header is not None:
            col = s.get("col")
            if isinstance(col, int):
                if not (0 <= col < len(header)):
                    problems.append("%s: 列序号 %d 越界（表头共 %d 列）" % (tag, col, len(header)))
            elif col not in header and col != "":
                problems.append("%s: 列名 %r 在表头中不存在" % (tag, col))
        if "delay_ms" in s:
            try:
                if int(s["delay_ms"]) < 0:
                    problems.append("%s: delay_ms 不能为负" % tag)
            except (TypeError, ValueError):
                problems.append("%s: delay_ms 必须是整数" % tag)

    if cfg.get("confirm") not in (None, "enter", "none"):
        problems.append("confirm 必须是 enter 或 none，当前 %r" % cfg.get("confirm"))
    rs, re_ = int(cfg.get("row_start", 1) or 1), int(cfg.get("row_end", 0) or 0)
    if rs < 1:
        problems.append("row_start 必须 ≥ 1")
    if re_ < 0:
        problems.append("row_end 不能为负（0=到末尾）")
    # 只对 button 类型做存在性校验（其它类型无需 header 也能部分校验）
    return problems


# ---------------- 主流程 ----------------

def ask_home_start(link):
    input(">>> 请把工控机上的鼠标光标移到『屏幕左上角』，然后回到这里按回车…")
    link.send({"op": "home"})


def _do_confirm(confirm_fn, prompt):
    """等待人工确认。confirm_fn 由 GUI 提供（通知主线程弹按钮）；
    否则用命令行 input()。用户输入 q / 取消时抛出 UserQuit 保留断点。"""
    if confirm_fn is not None:
        if not confirm_fn(prompt):
            raise UserQuit(prompt)
        return
    while True:
        ans = input(prompt)
        if ans.strip().lower() == "q":
            raise UserQuit(prompt)
        break


def run_once(link, cfg, records, start_row, dry_run=False, confirm_fn=None, header=None):
    steps = expand_steps(cfg)
    step_delay = int(cfg.get("step_delay_ms", 300))
    total = len(records)
    # 填充范围：row_start/row_end（1-based 数据行号），并与断点 start_row 取交集
    sel_start = int(cfg.get("row_start", 1) or 1)
    sel_end = int(cfg.get("row_end", 0) or 0)
    begin = max(start_row, sel_start - 1)
    end = total if sel_end <= 0 else min(total, sel_end)
    if begin >= end:
        print("没有待处理的行（已超出所选范围或已完成）。")
        return
    print("本次将填充数据行：第 %d–%d 行（表格共 %d 行数据）。" % (begin + 1, end, total))
    digest = _cfg_digest(cfg)
    confirm_mode = cfg.get("confirm", "enter")

    for i in range(begin, end):
        rec = records[i]
        cmds, summary = build_record_commands(steps, rec, step_delay, header)
        print("\n── 第 %d/%d 行 ──" % (i + 1, total))
        for name, val in summary.items():
            print("  %-12s %s" % (name, val))
        if dry_run:
            for c in cmds:
                print("    ", json.dumps(c, ensure_ascii=False))
            append_audit({"row": i + 1, "total": total, "data_file": cfg["data_file"],
                          "summary": summary, "confirm_required": False,
                          "confirmed": True, "status": "dry_run",
                          "time": time.strftime("%Y-%m-%d %H:%M:%S"), "run": "dry"})
            continue
        for c in cmds:
            if c["op"] == "__confirm__":
                _do_confirm(confirm_fn, c.get("label", "确认"))
                continue
            if c["op"] == "sleep":
                time.sleep(c["ms"] / 1000.0)
                continue
            resp = link.send(c)
            if resp.get("ack") == "error":
                raise RuntimeError("Pico 执行失败: %s （指令: %s）" % (resp.get("msg"), c))
        # 先人工确认，确认通过后才把这一行记为“已完成”并持久化断点
        # （避免用户发现该行错误退出后，断点把未确认的行永久跳过——漏填/错值）。
        confirmed = True
        if confirm_mode == "enter":
            _do_confirm(confirm_fn, ">>> 请到工控机屏幕上确认本行数值无误，再继续（q 退出）…")
        save_progress(i + 1, total, cfg["data_file"], digest=digest)
        append_audit({"row": i + 1, "total": total, "data_file": cfg["data_file"],
                      "summary": summary, "confirm_required": confirm_mode == "enter",
                      "confirmed": confirmed, "status": "filled",
                      "time": time.strftime("%Y-%m-%d %H:%M:%S"), "run": "cli"})
    print("\n全部完成 ✅  本次共 %d 行。" % (end - begin))


def iter_target_steps(steps):
    """返回流程里需要鼠标定位的步骤（click/dblclick/move/press_at）。"""
    for s in steps:
        if s.get("type") in ("click", "dblclick", "move", "press_at"):
            yield s


def iter_target_steps_flat(cfg):
    """遍历主流程和所有自定义按钮里需要鼠标定位的步骤。
    返回 (所属对象标签, 步骤dict)；步骤是 cfg 里的原始引用，
    校准修改坐标时会直接改到对应容器列表里，不会破坏 button 引用结构。"""
    groups = [("主流程", cfg.get("steps") or [])]
    for bname, bsteps in (cfg.get("buttons") or {}).items():
        groups.append(("按钮：" + bname, bsteps))
    for label, slist in groups:
        for s in slist:
            if s.get("type") in ("click", "dblclick", "move", "press_at"):
                yield label, s


def calibrate(link, cfg, cfg_path):
    """校准模式：逐个把光标移到流程里每个需要定位的步骤坐标，人工确认/微调，写回 config。"""
    print("== 坐标校准 ==")
    print("请确认工控机软件已打开并固定在平时工作的窗口位置。")
    ask_home_start(link)
    cal_targets = list(iter_target_steps_flat(cfg))
    if not cal_targets:
        print("当前流程和按钮里没有需要鼠标定位的步骤（全是键盘/输入），无需校准坐标。")
        return
    for idx, (label, t) in enumerate(cal_targets):
        print("\n[%d/%d] (%s) 步骤: %s  动作: %s  坐标 (%s, %s)"
              % (idx + 1, len(cal_targets), label, t.get("name") or t.get("type"), t["type"],
                 t.get("x"), t.get("y")))
        while True:
            link.send({"op": "move_to", "x": int(t["x"]), "y": int(t["y"])})
            time.sleep(0.3)
            ans = input("光标是否已对准目标位置？[y=对准 / n=微调 / q=退出] ").strip().lower()
            if ans == "y":
                break
            if ans == "q":
                print("已退出校准，未保存本次改动。")
                return
            if ans == "n":
                adj = input("输入像素偏移（如 +20,0 表示右移20，0,-10 表示上移10）: ").strip()
                try:
                    dx, dy = [int(s) for s in adj.replace(" ", "").split(",")]
                except ValueError:
                    print("格式不对，示例：+20,0 或 0,-10")
                    continue
                t["x"] = int(t["x"]) + dx
                t["y"] = int(t["y"]) + dy
                print("新坐标: (%d, %d)" % (t["x"], t["y"]))
    # 注意：不能用 expand_steps 的结果回写 steps（会把 button 展开而破坏结构），
    # 上面已直接修改 cfg 里的原始步骤对象，这里整体保存即可。
    with open(cfg_path, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)
    print("\n校准完成，坐标已写入 %s ✅" % cfg_path)


def main():
    parser = argparse.ArgumentParser(description="工控机自动填入数值工具（经 Pico HID 注入）")
    parser.add_argument("--config", default="config.json", help="配置文件路径（默认 config.json）")
    parser.add_argument("--calibrate", action="store_true", help="坐标校准模式")
    parser.add_argument("--dry-run", action="store_true", help="只打印指令序列，不连串口")
    parser.add_argument("--resume", action="store_true", help="直接从断点继续，不询问")
    args = parser.parse_args()

    cfg_path = args.config
    if not os.path.isabs(cfg_path):
        cfg_path = os.path.join(HERE, cfg_path)
    if not os.path.exists(cfg_path):
        sys.exit("找不到配置文件: %s\n请先执行: copy config.example.json config.json 然后编辑它" % cfg_path)
    cfg = read_config(cfg_path)

    if args.calibrate:
        link = Link(cfg["com_port"], cfg.get("baudrate", 115200))
        try:
            calibrate(link, cfg, cfg_path)
        finally:
            link.close()
        return

    try:
        header, records = load_records(cfg)
    except (FileNotFoundError, ValueError, ImportError, KeyError) as e:
        sys.exit("读取表格失败: %s" % e)
    print("读取表格: %s   共 %d 行数据" % (cfg["data_file"], len(records)))
    print("表头: %s" % ", ".join(h for h in header))

    problems = validate_config(cfg, header=header)
    if problems:
        sys.exit("配置校验未通过：\n- " + "\n- ".join(problems))

    # 列引用校验：列名必须在表头；整数列号必须在范围内（修复 CLI 对 col:2 整数列误报）
    missing = []
    for s in expand_steps(cfg):
        if s.get("type") == "input" and s.get("col") not in (None, ""):
            col = s["col"]
            if isinstance(col, int):
                if not (0 <= col < len(header)):
                    missing.append(col)
            elif col not in header:
                missing.append(col)
    if missing:
        sys.exit("步骤引用的表格列在表头中不存在: %s（表头: %s）" % (", ".join(str(m) for m in sorted(set(missing))),
                                                              ", ".join(header)))

    link = None
    try:
        start_row = 0
        prog = load_progress()
        if prog and prog.get("data_file") == cfg["data_file"] and prog.get("row", 0) < len(records):
            drifted = progress_is_drifted(prog, cfg)
            if args.resume:
                if drifted:
                    print("警告：检测到配置/表格与上次断点不一致（步骤、范围、表头或数据文件已变更）。")
                    print("     继续将从第 %d 行沿用断点，但可能与新配置错位。" % (prog["row"] + 1))
                start_row = prog["row"]
                print("从断点继续：第 %d 行之后。" % (start_row + 1))
            else:
                tail = "（且配置已变更，继续可能错位！）" if drifted else ""
                ans = input("存在未完成的进度（已完成 %d/%d 行）%s，从断点继续？[Y/n] "
                            % (prog["row"], prog.get("total", len(records)), tail)).strip().lower()
                if ans not in ("n", "no"):
                    start_row = prog["row"]

        if args.dry_run:
            print("\n== DRY RUN：仅打印指令，不发送 ==")
            run_once(None, cfg, records, start_row, dry_run=True, header=header)
            return

        link = Link(cfg["com_port"], cfg.get("baudrate", 115200))
        print("已连接串口 %s，检测 Pico…" % cfg["com_port"])
        if not link.ping():
            sys.exit("Pico 无响应。检查：串口号是否对、TX/RX 是否接反、Pico 固件是否已烧录。")
        print("Pico 在线 ✅")

        if cfg.get("start_home", True):
            ask_home_start(link)
        else:
            print("注意：start_home=false，请确保工控机鼠标当前就在屏幕左上角，或已人工归零。")

        try:
            run_once(link, cfg, records, start_row, header=header)
        except UserQuit:
            print("\n已中断。进度已保存，下次运行可用 --resume 继续。")
            return
        except (KeyboardInterrupt, RuntimeError, TimeoutError) as e:
            print("\n出错/中断: %s" % e)
            print("进度已保存，修复后可用 --resume 继续。")
            return
        clear_progress()
    finally:
        if link:
            link.close()


if __name__ == "__main__":
    main()